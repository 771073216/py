import os
import re
import shutil
import sqlite3

from PySide6.QtCore import QThread, Signal, QDate, QTime
from PySide6.QtGui import QGuiApplication
from PySide6.QtWidgets import QMessageBox, QInputDialog, QFileDialog, QMainWindow, QHeaderView, QApplication, QDialog
from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Side, Border, colors, Font, Alignment
from openpyxl.workbook.workbook import Workbook
from requests.api import get
from win32com.client import Dispatch
from win32com.client.gencache import EnsureDispatch
from win32com.universal import com_error
from win32print import GetDefaultPrinter, EnumPrinters

from need.dialog import *
from need.mainwindow import *

desktop_path = os.path.join(os.path.expanduser('~'), "Desktop")
sql_dir = os.path.join(os.path.expanduser('~'), "sql")
sql_file = os.path.join(os.path.expanduser(sql_dir), "list.db")
backup_dir = os.path.join(os.path.expanduser(sql_dir), "backup")
update_file = os.getenv('TEMP') + '/update.exe'
temp_excel = os.getenv('TEMP') + '/prepare_print.xlsx'

alignment_center = Alignment(horizontal='center', vertical='center')
border_set = Border(left=Side(style='thin', color=colors.BLACK),
                    right=Side(style='thin', color=colors.BLACK),
                    top=Side(style='thin', color=colors.BLACK),
                    bottom=Side(style='thin', color=colors.BLACK))

styles = '''     
QCalendarWidget QWidget#qt_calendar_navigationbar {background-color: rgb(255, 255, 255);}
QToolButton#qt_calendar_monthbutton,#qt_calendar_yearbutton {color: rgb(0, 0, 0);font: 微软雅黑;}
QCalendarWidget QToolButton#qt_calendar_prevmonth{qproperty-icon: url(left.png);}
QCalendarWidget QToolButton#qt_calendar_nextmonth{qproperty-icon: url(right.png);}
QCalendarWidget QTableView {alternate-background-color: rgb(245, 245, 245);background-color: rgb(245, 245, 245);}
'''

version = 'v1.0'


def fix_missing():
    conn = sqlite3.connect(sql_file)
    c = conn.cursor()
    sql = "select name from sqlite_master where type='table'"
    result = c.execute(sql).fetchall()
    if len(result) != 6:
        add_sql_1()
        add_sql_2()
        add_sql_3()
        add_sql_4()
        add_sql_5()
        add_sql_6()
        conn.close()


def backup():
    is_backup_dir = os.path.exists(backup_dir)
    if not is_backup_dir:
        os.makedirs(backup_dir)
    date1 = QDate.currentDate().toString(Qt.ISODate)
    ppdir = backup_dir + '/' + date1
    isdir = os.path.exists(ppdir)
    if not isdir:
        os.makedirs(ppdir)
        shutil.copyfile(sql_file, ppdir + '/' + 'list.db')


def check_sql():
    isdir = os.path.exists(sql_dir)
    if not isdir:
        os.makedirs(sql_dir)
        conn = sqlite3.connect(sql_file)
        add_sql_1()
        add_sql_2()
        add_sql_3()
        add_sql_4()
        add_sql_5()
        add_sql_6()
        conn.close()


def add_sql_1():
    try:
        conn = sqlite3.connect(sql_file)
        c = conn.cursor()
        c.execute('''CREATE TABLE COMPANY
                        (ID            INT     NOT NULL,
                           NAME           TEXT    DEFAULT NULL,
                           REMINDER       INT     DEFAULT NULL,
                           MODEL          TEXT    DEFAULT NULL,
                           UNIT           TEXT    DEFAULT NULL,
                           COST           INT     DEFAULT NULL,
                           PRICE          INT     DEFAULT NULL,
                           PRIMARY KEY (`ID`));''')
        conn.commit()
    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
        pass


def add_sql_2():
    try:
        conn = sqlite3.connect(sql_file)
        c = conn.cursor()
        c.execute('''CREATE TABLE SALE
                            (DATE             TEXT    NOT NULL,
                               ID             INT     NOT NULL,
                               NAME           TEXT    DEFAULT NULL,
                               MODEL          TEXT    DEFAULT NULL,
                               UNIT           TEXT    DEFAULT NULL,
                               COST           INT     DEFAULT NULL,
                               PRICE          INT     DEFAULT NULL,
                               NUM            INT     DEFAULT NULL,
                               COST_SUM       INT     DEFAULT NULL,
                               SUM            INT     DEFAULT NULL,
                               GET            INT     DEFAULT NULL);''')
        conn.commit()
    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
        pass


def add_sql_3():
    try:
        conn = sqlite3.connect(sql_file)
        c = conn.cursor()
        c.execute('''CREATE TABLE SELL
                                    (ID             INT     NOT NULL,
                                       NAME           TEXT    DEFAULT NULL,
                                       MODEL          TEXT    DEFAULT NULL,
                                       UNIT           TEXT    DEFAULT NULL,
                                       NUM            INT     DEFAULT NULL,
                                       PRICE          INT     DEFAULT NULL,
                                       SUM            INT     DEFAULT NULL,
                                       PRIMARY KEY (`ID`));''')
        conn.commit()
    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
        pass


def add_sql_4():
    try:
        conn = sqlite3.connect(sql_file)
        c = conn.cursor()
        c.execute('''CREATE TABLE GOODS
                              (DATE           TEXT    NOT NULL,
                               ID             INT     NOT NULL,
                               NAME           TEXT    DEFAULT NULL,
                               MODEL          TEXT    DEFAULT NULL,
                               UNIT           TEXT    DEFAULT NULL,
                               COST           INT     DEFAULT NULL,
                               NUM            INT     DEFAULT NULL,
                               COST_SUM       INT     DEFAULT NULL);''')
        conn.commit()
    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
        pass


def add_sql_5():
    try:
        conn = sqlite3.connect(sql_file)
        c = conn.cursor()
        c.execute('''CREATE TABLE INFO
                                      (NAME           TEXT    NOT NULL,
                                       LOCALE         TEXT    NOT NULL,
                                       NUMBER         TEXT    NOT NULL,
                                       PRINTER        TEXT    NOT NULL);''')
        conn.commit()
        c.execute("INSERT INTO INFO (NAME,LOCALE,NUMBER,PRINTER) VALUES ('', '', '', '')")
        conn.commit()
    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
        pass


def add_sql_6():
    try:
        conn = sqlite3.connect(sql_file)
        c = conn.cursor()
        c.execute('''CREATE TABLE CAL
                                              (A1         INT    NOT NULL,
                                               A2         INT    NOT NULL,
                                               A3         INT    NOT NULL,
                                               A4         INT    NOT NULL,
                                               A5         INT    NOT NULL,
                                               A6         INT    NOT NULL,
                                               A7         INT    NOT NULL,
                                               B1         INT    NOT NULL,
                                               B2         INT    NOT NULL,
                                               B3         INT    NOT NULL,
                                               B4         INT    NOT NULL,
                                               B5         INT    NOT NULL,
                                               B6         INT    NOT NULL,
                                               B7         INT    NOT NULL,
                                               B8         INT    NOT NULL,
                                               B9         INT    NOT NULL,
                                               B10        INT    NOT NULL,
                                               B11        INT    NOT NULL,
                                               B12        INT    NOT NULL);''')
        conn.commit()
        c.execute("INSERT INTO CAL (A1,A2,A3,A4,A5,A6,A7,B1,B2,B3,B4,B5,B6,B7,B8,B9,B10,B11,B12) "
                  "VALUES ('', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '')")
        conn.commit()
    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
        pass


class FirstWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super(FirstWindow, self).__init__()
        self.result = None
        self.setupUi(self)
        check_sql()
        backup()
        fix_missing()
        self.up = Update()
        self.download_update = DownloadUpdate()
        self.row1 = None
        self.col1 = None

        self.put_goods_data()
        self.put_sale_data()
        self.put_stock_data()
        self.put_prepare_goods_data()
        self.put_sell_out_data()
        self.put_setting_data()
        self.put_calculate_data()
        self.put_printer_data()

        self.table_goods.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_goods.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_goods.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.table_goods.cellPressed.connect(self.get_click)
        self.table_goods.verticalHeader().setVisible(False)

        self.table_new_goods.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)

        self.table_sale.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_sale.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_sale.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table_sale.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table_sale.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.table_sale.horizontalHeader().setSectionResizeMode(7, QHeaderView.ResizeMode.ResizeToContents)
        self.table_sale.verticalHeader().setVisible(False)
        self.table_sale.cellPressed.connect(self.get_sale_click)

        self.table_stock.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_stock.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_stock.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.table_stock.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.table_stock.horizontalHeader().setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        self.table_stock.horizontalHeader().setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)
        self.table_stock.verticalHeader().setVisible(False)
        self.table_stock.cellPressed.connect(self.get_stock_click)

        self.table_prepare_goods.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_prepare_goods.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.ResizeToContents)
        self.table_prepare_goods.horizontalHeader().setSectionResizeMode(
            4, QHeaderView.ResizeMode.ResizeToContents)
        self.table_prepare_goods.cellPressed.connect(self.get_prepare_goods_click)
        self.table_prepare_goods.verticalHeader().setVisible(False)

        self.table_sell_out.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table_sell_out.horizontalHeader().setSectionResizeMode(0,
                                                                    QHeaderView.ResizeMode.ResizeToContents)
        self.table_sell_out.horizontalHeader().setSectionResizeMode(4,
                                                                    QHeaderView.ResizeMode.ResizeToContents)
        self.table_sell_out.cellPressed.connect(self.get_sell_out_click)
        self.table_sell_out.verticalHeader().setVisible(False)

        self.btn_new_goods_1.clicked.connect(lambda: self.table_new_goods.insertRow(self.table_new_goods.rowCount()))
        self.btn_new_goods_2.clicked.connect(
            lambda: self.table_new_goods.removeRow(self.table_new_goods.rowCount() - 1))
        self.btn_new_goods_3.clicked.connect(self.merge_new_goods_table)

        self.btn_goods_1.clicked.connect(lambda: self.search_goods_by('name'))
        self.btn_goods_2.clicked.connect(lambda: self.search_goods_by('id'))
        self.btn_goods_3.clicked.connect(self.clean_goods_search)
        self.btn_goods_4.clicked.connect(self.del_goods_row)
        self.btn_goods_5.clicked.connect(self.update_goods_reminder)

        self.btn_sale_1.clicked.connect(lambda: self.search_sale_by('name'))
        self.btn_sale_2.clicked.connect(lambda: self.search_sale_by('date'))
        self.btn_sale_3.clicked.connect(self.clean_sale_search)
        self.btn_sale_4.clicked.connect(self.sale_rollback)
        self.btn_sale_5.clicked.connect(self.del_sale_row)

        self.btn_stock_1.clicked.connect(lambda: self.search_stock_by('name'))
        self.btn_stock_2.clicked.connect(lambda: self.search_stock_by('date'))
        self.btn_stock_3.clicked.connect(self.clean_stock_search)
        self.btn_stock_4.clicked.connect(self.del_stock_row)

        self.btn_sell_1.clicked.connect(self.sell_goods)
        self.btn_sell_2.clicked.connect(self.del_sell_out_row)
        self.btn_sell_3.clicked.connect(self.sell_out)
        self.btn_sell_4.clicked.connect(lambda: self.search_prepare_goods_by('name'))
        self.btn_sell_5.clicked.connect(lambda: self.search_prepare_goods_by('id'))
        self.btn_sell_6.clicked.connect(self.clean_prepare_goods_search)

        self.btn_settings_1.clicked.connect(self.settings)

        self.pushButton.clicked.connect(self.get_update)
        self.label_5.setText(self.label_5.text() + version)

        self.action.triggered.connect(self.excel_template)
        self.action_Excel.triggered.connect(self.read_excel)
        self.action_Excel_2.triggered.connect(self.save_excel)
        self.action_Excel_3.triggered.connect(self.save_sellout_excel_by_time)
        self.action_Excel_4.triggered.connect(self.save_sellout_excel)

        self.text_cal_1.textEdited.connect(self.calculate)
        self.text_cal_2.textEdited.connect(self.calculate)

        self.text_a1.clicked.connect(lambda: self.update_calculate_data('a1'))
        self.text_a2.clicked.connect(lambda: self.update_calculate_data('a2'))
        self.text_a3.clicked.connect(lambda: self.update_calculate_data('a3'))
        self.text_a4.clicked.connect(lambda: self.update_calculate_data('a4'))
        self.text_a5.clicked.connect(lambda: self.update_calculate_data('a5'))
        self.text_a6.clicked.connect(lambda: self.update_calculate_data('a6'))
        self.text_a7.clicked.connect(lambda: self.update_calculate_data('a7'))

        self.text_b1.clicked.connect(lambda: self.update_calculate_data('b1'))
        self.text_b2.clicked.connect(lambda: self.update_calculate_data('b2'))
        self.text_b3.clicked.connect(lambda: self.update_calculate_data('b3'))
        self.text_b4.clicked.connect(lambda: self.update_calculate_data('b4'))
        self.text_b5.clicked.connect(lambda: self.update_calculate_data('b5'))
        self.text_b6.clicked.connect(lambda: self.update_calculate_data('b6'))
        self.text_b7.clicked.connect(lambda: self.update_calculate_data('b7'))
        self.text_b8.clicked.connect(lambda: self.update_calculate_data('b8'))
        self.text_b9.clicked.connect(lambda: self.update_calculate_data('b9'))
        self.text_b10.clicked.connect(lambda: self.update_calculate_data('b10'))
        self.text_b11.clicked.connect(lambda: self.update_calculate_data('b11'))
        self.text_b12.clicked.connect(lambda: self.update_calculate_data('b12'))

        self.pushButton_2.clicked.connect(self.update_printer)

        self.dateEdit.setDate(QDate.currentDate())
        self.dateEdit.calendarWidget().setStyleSheet(styles)
        self.dateEdit_2.setDate(QDate.currentDate())
        self.dateEdit_2.calendarWidget().setStyleSheet(styles)

    def exec_sqlite(self, e, sql):
        conn = sqlite3.connect(sql_file)
        if e == 'r':
            results = conn.cursor().execute(sql).fetchall()
            conn.close()
            return results
        elif e == 'w':
            try:
                conn.cursor().execute(sql)
                conn.commit()
            except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                conn.rollback()
                QMessageBox.warning(self, '警告！', '数据库操作失败！')
            except sqlite3.IntegrityError:
                return 0
            else:
                return 1
            finally:
                conn.close()

    def merge_new_goods_table(self):
        row = self.table_new_goods.rowCount()
        if row == 0:
            return
        value = []
        error_flag = 0
        merge_flag = 0
        err = []
        values = []
        for i in range(row):
            for m in range(7):
                try:
                    value.append(self.table_new_goods.item(i, m).text())
                except AttributeError:
                    error_flag = 1
                    break
            if value.count('') == 0 and value and len(value) == 7:
                sql = "INSERT INTO COMPANY (ID,NAME,REMINDER,MODEL,UNIT,COST,PRICE) \
                                           VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
                      (value[0], value[1], value[2], value[3], value[4], value[5], value[6])
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    merge_flag = 1
                    values.append(i)
                elif status == 0:
                    err.append('id_err')
                    error_flag = 1
            else:
                err.append('value_err')
            value = []
        self.clean_goods_search()
        self.clean_prepare_goods_search()
        for i in reversed(values):
            self.table_new_goods.removeRow(i)
        k = 0
        for i in err:
            k = k + 1
            if i == 'id_err':
                QMessageBox.warning(self, '警告！', '合并失败，第 ' + str(k) + ' 行商品id冲突！')
            else:
                QMessageBox.warning(self, '警告！', '合并失败，第 ' + str(k) + ' 行信息有错误！')
        if merge_flag == 1 and error_flag == 1:
            QMessageBox.information(self, '提示！', '正确的数据合并成功！')
        elif merge_flag == 1 and error_flag == 0:
            QMessageBox.information(self, '提示！', '合并成功！')

    def get_click(self, row, col):
        self.table_goods.item(row, col).text()
        self.row1 = row

    def click(self, q):
        pp = self.table_goods.item(self.table_goods.currentRow(), q).text()
        return pp

    def del_goods_row(self):
        if not self.row1 and self.row1 != 0:
            QMessageBox.warning(self, '提示！', '请点击要删除的项目！')
        else:
            reply = QMessageBox.question(self, '警告', '是否删除所选数据？')
            if reply == QMessageBox.Yes:
                del_id = self.click(0)
                sql = "DELETE from COMPANY where ID='%s'" % del_id
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    self.table_goods.removeRow(self.table_goods.currentRow())
                    QMessageBox.information(self, '提示', '删除成功！')
            self.row1 = None

    def update_goods_reminder(self):
        if not self.row1 and self.row1 != 0:
            QMessageBox.warning(self, '警告！', '请点击要更新的项目！')
        else:
            id1 = self.click(0)
            name = self.click(1)
            reminder = self.click(2)
            model = self.click(3)
            unit = self.click(4)
            cost = self.click(5)
            ok = bool()
            num, ok = QInputDialog.getInt(self, '输入框', '输入数字', ok)
            if ok is False:
                self.row1 = None
                return
            elif ok is True and num <= 0:
                QMessageBox.information(self, '提示！', '请输入一个大于 0 的数！')
            else:
                after_num = num + int(reminder)
                cost_sum = num * int(cost)
                sql = "UPDATE COMPANY set REMINDER = '%s' where ID='%s'" % \
                      (after_num, self.click(0))
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    self.table_goods.item(int(self.row1), 2).setText(str(after_num))
                date1 = QDate.currentDate().toString(Qt.ISODate) + ' ' + QTime.currentTime().toString(Qt.ISODate)
                sql1 = "INSERT INTO GOODS (DATE,ID,NAME,MODEL,UNIT,COST,NUM,COST_SUM) \
                                           VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
                       (date1, id1, name, model, unit, cost, num, cost_sum)
                status = self.exec_sqlite('w', sql1)
                if status == 1:
                    self.clean_prepare_goods_search()
                    self.clean_stock_search()
                    QMessageBox.information(self, '提示！', '更新成功！')
                    self.row1 = None

    def clean_goods_search(self):
        self.text_1.clear()
        self.put_goods_data()

    def search_goods_by(self, var):
        self.row1 = None
        text = self.text_1.text()
        if not text:
            self.put_goods_data()
        else:
            sql = "SELECT * FROM COMPANY WHERE " + var + " like '%" + text + "%'"
            results = self.exec_sqlite('r', sql)
            self.write_goods_data(results)

    def put_goods_data(self):
        sql = "select * from COMPANY order by id"
        results = self.exec_sqlite('r', sql)
        self.write_goods_data(results)

    def write_goods_data(self, results):
        for i in range(self.table_goods.rowCount()):
            self.table_goods.removeRow(0)
        i = 0
        self.table_goods.setRowCount(len(results))
        for row in results:
            for m in range(7):
                self.table_goods.setItem(i, m, QTableWidgetItem())
                item_add = self.table_goods.item(i, m)
                item_add.setText(str(row[m]))
                item_add.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            i = i + 1

    def write_sale_data(self, results):
        for i in range(self.table_sale.rowCount()):
            self.table_sale.removeRow(0)
        i = 0
        self.table_sale.setRowCount(len(results))
        for row in results:
            for m in range(11):
                self.table_sale.setItem(i, m, QTableWidgetItem())
                item_add = self.table_sale.item(i, m)
                item_add.setText(str(row[m]))
                item_add.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            i = i + 1

    def put_sale_data(self):
        sql = "SELECT *  from SALE"
        results = self.exec_sqlite('r', sql)
        self.write_sale_data(results)

    def clean_sale_search(self):
        self.text_sale_1.clear()
        self.put_sale_data()

    def search_sale_by(self, var):
        self.row1 = None
        if var == 'date':
            text = self.dateEdit_2.date().toString(Qt.ISODate)
        else:
            text = self.text_sale_1.text()
        if not text:
            self.put_sale_data()
        else:
            sql = "SELECT * FROM SALE WHERE " + var + " like '%" + text + "%'"
            results = self.exec_sqlite('r', sql)
            self.write_sale_data(results)

    def get_sale_click(self, row, col):
        self.table_sale.item(row, col).text()
        self.row1 = row

    def sale_click(self, q):
        pp = self.table_sale.item(self.table_sale.currentRow(), q).text()
        return pp

    def del_sale_row(self):
        if not self.row1 and self.row1 != 0:
            QMessageBox.warning(self, '提示！', '请点击要删除的项目！')
        else:
            reply = QMessageBox.question(self, '警告', '是否删除所选数据？')
            if reply == QMessageBox.Yes:
                del_date = self.sale_click(0)
                del_id = self.sale_click(1)
                sql = "DELETE from SALE where DATE='%s' and ID='%s'" % (del_date, del_id)
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    self.table_sale.removeRow(self.table_sale.currentRow())
                    QMessageBox.information(self, '提示', '删除成功！')
            self.row1 = None

    def sale_rollback(self):
        if not self.row1 and self.row1 != 0:
            QMessageBox.warning(self, '提示！', '请点击要退货的项目！')
        else:
            ok = bool()
            num, ok = QInputDialog.getInt(self, '退货数量', '输入数字', ok)
            if ok is False:
                self.row1 = None
                return
            elif ok is True and num <= 0:
                QMessageBox.information(self, '提示！', '请输入一个大于 0 的数！')
            else:
                sql = "SELECT REMINDER from COMPANY  where ID='%s' and NAME='%s' and MODEL='%s' and COST='%s' and " \
                      "PRICE='%s'" % (self.sale_click(1), self.sale_click(2), self.sale_click(3), self.sale_click(5),
                                      self.sale_click(6))
                try:
                    self.exec_sqlite('r', sql)[0]
                except IndexError:
                    QMessageBox.warning(self, '提示！', '退货失败，数据库没有相同商品！\n'
                                                     '注：商品id、商品名称、商品型号、成本价和单价需要同时匹配。')
                    return
                sql = "SELECT * FROM SALE WHERE date = '%s' and id = '%s'" % (self.sale_click(0), self.sale_click(1))
                results = self.exec_sqlite('r', sql)[0]
                after_num = int(results[7]) - int(num)
                if after_num == 0:
                    sql = "DELETE from SALE where DATE='%s' and ID='%s'" % (self.sale_click(0), self.sale_click(1))
                    status = self.exec_sqlite('w', sql)
                    if status == 1:
                        self.table_sale.removeRow(self.table_sale.currentRow())
                elif after_num < 0:
                    QMessageBox.warning(self, '警告！', '更新失败，退货数量不能大于售出数量！')
                    return
                else:
                    after_cost_sum = after_num * int(self.sale_click(5))
                    after_price_sum = after_num * int(self.sale_click(6))
                    after_get = after_price_sum - after_cost_sum
                    sql = "UPDATE SALE set NUM='%s', COST_SUM='%s',SUM='%s',GET='%s' where DATE = '%s' and ID = '%s'" \
                          % (after_num, after_cost_sum, after_price_sum, after_get,
                             self.sale_click(0), self.sale_click(1))
                    status = self.exec_sqlite('w', sql)
                    if status == 1:
                        self.table_sale.item(int(self.row1), 7).setText(str(after_num))
                        self.table_sale.item(int(self.row1), 8).setText(str(after_cost_sum))
                        self.table_sale.item(int(self.row1), 9).setText(str(after_price_sum))
                        self.table_sale.item(int(self.row1), 10).setText(str(after_get))
                sql = "SELECT REMINDER from COMPANY  where ID='%s'" % self.sale_click(1)
                results = self.exec_sqlite('r', sql)[0]
                goods_reminder = int(results[0]) + int(num)
                sql = "UPDATE COMPANY set REMINDER = '%s' where ID='%s'" % \
                      (goods_reminder, self.sale_click(1))
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    self.clean_goods_search()
                    self.clean_prepare_goods_search()
                    QMessageBox.information(self, '提示！', '更新成功！')
                self.row1 = None

    def write_stock_data(self, results):
        for i in range(self.table_stock.rowCount()):
            self.table_stock.removeRow(0)
        i = 0
        self.table_stock.setRowCount(len(results))
        for row in results:
            for m in range(8):
                self.table_stock.setItem(i, m, QTableWidgetItem())
                item_add = self.table_stock.item(i, m)
                item_add.setText(str(row[m]))
                item_add.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            i = i + 1

    def put_stock_data(self):
        sql = "SELECT *  from GOODS"
        results = self.exec_sqlite('r', sql)
        self.write_stock_data(results)

    def clean_stock_search(self):
        self.text_stock_1.clear()
        self.put_stock_data()

    def search_stock_by(self, var):
        self.row1 = None
        if var == 'date':
            text = self.dateEdit.date().toString(Qt.ISODate)
        else:
            text = self.text_stock_1.text()
        if not text:
            self.put_stock_data()
        else:
            sql = "SELECT * FROM GOODS WHERE " + var + " like '%" + text + "%'"
            results = self.exec_sqlite('r', sql)
            self.write_stock_data(results)

    def get_stock_click(self, row, col):
        self.table_stock.item(row, col).text()
        self.row1 = row

    def stock_click(self, q):
        pp = self.table_stock.item(self.table_stock.currentRow(), q).text()
        return pp

    def del_stock_row(self):
        if not self.row1 and self.row1 != 0:
            QMessageBox.warning(self, '提示！', '请点击要删除的项目！')
        else:
            reply = QMessageBox.question(self, '警告', '是否删除所选数据？')
            if reply == QMessageBox.Yes:
                del_date = self.stock_click(0)
                del_id = self.stock_click(1)
                sql = "DELETE from GOODS where DATE='%s' and ID='%s'" % (del_date, del_id)
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    self.table_stock.removeRow(self.table_stock.currentRow())
                    QMessageBox.information(self, '提示', '删除成功！')
            self.row1 = None

    def write_prepare_goods_data(self, results):
        for i in range(self.table_prepare_goods.rowCount()):
            self.table_prepare_goods.removeRow(0)
        i = 0
        self.table_prepare_goods.setRowCount(len(results))
        for row in results:
            rows = [row[0], row[1], row[2], row[3], row[4], row[6]]
            for m in range(6):
                self.table_prepare_goods.setItem(i, m, QTableWidgetItem())
                item_add = self.table_prepare_goods.item(i, m)
                item_add.setText(str(rows[m]))
                item_add.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            i = i + 1

    def put_prepare_goods_data(self):
        sql = "select * from COMPANY order by id"
        results = self.exec_sqlite('r', sql)
        self.write_prepare_goods_data(results)

    def clean_prepare_goods_search(self):
        self.text_sell.clear()
        self.put_prepare_goods_data()

    def search_prepare_goods_by(self, var):
        self.row1 = None
        text = self.text_sell.text()
        if not text:
            self.put_prepare_goods_data()
        else:
            sql = "SELECT * FROM COMPANY WHERE " + var + " like '%" + text + "%'"
            results = self.exec_sqlite('r', sql)
            self.write_prepare_goods_data(results)

    def get_prepare_goods_click(self, row, col):
        self.table_prepare_goods.item(row, col).text()
        self.row1 = row

    def prepare_goods_click(self, q):
        pp = self.table_prepare_goods.item(self.table_prepare_goods.currentRow(), q).text()
        return pp

    def sell_goods(self):
        if not self.row1 and self.row1 != 0:
            QMessageBox.warning(self, '警告！', '请点击要更新的项目！')
        else:
            id1 = self.prepare_goods_click(0)
            name = self.prepare_goods_click(1)
            reminder = self.prepare_goods_click(2)
            model = self.prepare_goods_click(3)
            unit = self.prepare_goods_click(4)
            price = self.prepare_goods_click(5)
            ok = bool()
            num, ok = QInputDialog.getInt(self, '输入框', '输入数字', ok)
            if ok is False:
                self.row1 = None
                return
            elif ok is True and num <= 0:
                QMessageBox.information(self, '提示！', '请输入一个大于 0 的数！')
            else:
                after_num = int(reminder) - num
                if after_num < 0:
                    QMessageBox.information(self, '提示！', '库存不能小于0！')
                    return
                get_sum = num * int(price)
                sql = "INSERT INTO SELL (ID,NAME,MODEL,UNIT,NUM,PRICE,SUM) \
                                           VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
                      (id1, name, model, unit, num, price, get_sum)
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    sql = "UPDATE COMPANY set REMINDER = '%s' where ID='%s'" % (after_num, self.prepare_goods_click(0))
                    status = self.exec_sqlite('w', sql)
                    if status == 1:
                        self.table_prepare_goods.item(int(self.row1), 2).setText(str(after_num))
                        self.put_goods_data()
                        self.put_sell_out_data()
                        QMessageBox.information(self, '提示！', '更新成功！')
                elif status == 0:
                    QMessageBox.warning(self, '警告！', '更新失败，请勿重复出售相同的商品！')
                self.row1 = None

    def put_sell_out_data(self):
        sql = "SELECT * FROM SELL"
        results = self.exec_sqlite('r', sql)
        for i in range(self.table_sell_out.rowCount()):
            self.table_sell_out.removeRow(0)
        i = 0
        for row in results:
            self.table_sell_out.setRowCount(len(results))
            for m in range(7):
                item = QTableWidgetItem()
                self.table_sell_out.setItem(i, m, item)
                item_add = self.table_sell_out.item(i, m)
                item_add.setText(str(row[m]))
                item_add.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            i = i + 1

    def get_sell_out_click(self, row, col):
        self.table_sell_out.item(row, col).text()
        self.row1 = row

    def sell_out_click(self, q):
        pp = self.table_sell_out.item(self.table_sell_out.currentRow(), q).text()
        return pp

    def del_sell_out_row(self):
        if not self.row1 and self.row1 != 0:
            QMessageBox.warning(self, '提示！', '请点击要删除的项目！')
        else:
            reply = QMessageBox.question(self, '警告', '是否删除所选数据？')
            if reply == QMessageBox.Yes:
                del_id = self.sell_out_click(0)
                sql = "DELETE from SELL where ID='%s'" % del_id
                status = self.exec_sqlite('w', sql)
                if status == 1:
                    sql = "SELECT REMINDER from COMPANY where ID='%s'" % del_id
                    results = self.exec_sqlite('r', sql)[0]
                    k = results[0] + int(self.sell_out_click(4))
                    sql = "UPDATE COMPANY set REMINDER = '%s' where ID='%s'" % (k, del_id)
                    self.exec_sqlite('w', sql)
                    self.clean_goods_search()
                    self.clean_prepare_goods_search()
                    self.table_sell_out.removeRow(self.table_sell_out.currentRow())
                    QMessageBox.information(self, '提示', '删除成功！')
            self.row1 = None

    def sell_out(self):
        if self.table_sell_out.rowCount() == 0:
            return
        sql = "SELECT *  from INFO"
        get_info = self.exec_sqlite('r', sql)[0]
        name = get_info[0]
        locale = get_info[1]
        number = get_info[2]
        if name == '' or locale == '' or number == '':
            QMessageBox.information(self, '提示！', '未填写Excel信息！')
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        ws.cell(1, 2).value = name
        ws.cell(1, 2).font = Font(size="16")
        ws.cell(2, 1).value = '客户：'
        ws.cell(2, 5).value = QDate.currentDate().toString('yyyy年MM月dd日')
        n = 0
        var = ['序号', '名称', '型号', '单位', '数量', '单价', '总价']
        for i in range(7):
            n = n + 1
            ws.cell(3, n).value = var[i]
            ws.cell(3, n).border = border_set
            ws.cell(3, n).alignment = alignment_center
        sql = "SELECT *  from SELL"
        results = self.exec_sqlite('r', sql)
        n = 3
        m = 0
        for row in results:
            rows = [row[1], row[2], row[3], row[4], row[5], row[6]]
            n = n + 1
            m = m + 1
            ws.cell(n, 1).value = m
            ws.cell(n, 1).border = border_set
            ws.cell(n, 1).alignment = alignment_center
            for i in range(len(rows)):
                try:
                    float(rows[i])
                except ValueError:
                    ws.cell(n, i + 2).value = rows[i]
                    ws.cell(n, i + 2).border = border_set
                    ws.cell(n, i + 2).alignment = alignment_center
                else:
                    ws.cell(n, i + 2).value = rows[i]
                    ws.cell(n, i + 2).border = border_set
                    ws.cell(n, i + 2).data_type = "int"
                    ws.cell(n, i + 2).alignment = alignment_center
        if n == 4:
            bbc = '=G4'
        else:
            bbc = '=SUM(G4:G' + str(n) + ')'
        mm = str(n + 1)
        bb = '=TEXT(SUBSTITUTE(IF(G' + mm + '="","",IF(MOD(G' + mm + ',1)=0,TEXT(INT(G' + mm + \
             '),"[DBNum2][$-804]G/通用格式元整"),(TEXT(INT(G' + mm + '),"[DBNum2][$-804]G/通用格式元")&TEXT((INT(G' + mm + \
             '*10)-INT(G' + mm + ')*10),"[DBNum2][$-804]G/通用格式角")&TEXT((INT(G' + mm + '*100)-INT(G' + mm + \
             '*10)*10),"[DBNum2][$-804]G/通用格式分")))),"零角","零"),)'
        var = [' ', '大写', bb, ' ', ' ', ' ', bbc]
        m = 0
        for i in range(7):
            m = m + 1
            ws.cell(n + 1, m).value = var[i]
            ws.cell(n + 1, m).border = border_set
            ws.cell(n + 1, m).alignment = alignment_center
        var = [locale, '电话：', number]
        for i in range(3):
            ws.cell(n + 3, i + 3).value = var[i]
        ws.column_dimensions['B'].width = 20.0
        ws.column_dimensions['C'].width = 20.0
        ws.merge_cells('E2:F2')
        ws.merge_cells('B1:C1')
        ws.merge_cells('B2:C2')
        merge_tel = 'E' + str(n + 3) + ':' + 'F' + str(n + 3)
        ws.merge_cells(merge_tel)
        if self.checkBox_1.isChecked():
            path, ok = QFileDialog.getSaveFileName(self, '保存文件', desktop_path + '/出货单.xlsx', "Excel (*.xlsx)")
            if ok == '':
                return
            try:
                wb.save(path)
            except PermissionError:
                QMessageBox.warning(self, '警告！', '导出失败，请检查文件是否被占用！')
            else:
                self.put_sale()
                self.clean_sale()
                QMessageBox.information(self, '提示！', '导出成功！')
            finally:
                wb.close()
        else:
            try:
                wb.save(temp_excel)
            except PermissionError:
                QMessageBox.warning(self, '警告！', '导出失败，请检查文件是否被占用！')
            else:
                try:
                    EnsureDispatch('Excel.Application')
                except com_error:
                    try:
                        EnsureDispatch('KET.Application')
                    except com_error:
                        QMessageBox.warning(self, '提示！', '打印失败！请安装 WPS 或者 Microsoft Office')
                        return
                    else:
                        o = Dispatch("KET.Application")
                        o.Visible = 0
                else:
                    o = Dispatch("Excel.Application")
                    o.Visible = 0
                wb = o.Workbooks.Open(r'出货单.xlsx')
                ws = wb.Worksheets[1]
                try:
                    ws.PrintOut(ActivePrinter=self.label_11.text())
                except com_error:
                    QMessageBox.information(self, '提示！', '打印失败！')
                else:
                    self.put_sale()
                    self.clean_sale()
                    QMessageBox.information(self, '提示！', '打印成功！')

    def put_sale(self):
        date = QDate.currentDate().toString(Qt.ISODate) + ' ' + QTime.currentTime().toString(Qt.ISODate)
        sql = "SELECT * FROM SELL"
        result = self.exec_sqlite('r', sql)
        for i in range(len(result)):
            value = result[i]
            sql = "select * from COMPANY where ID='%s'" % value[0]
            result1 = self.exec_sqlite('r', sql)[0]
            cost_sum = result1[5] * value[4]
            price = result1[6] * value[4]
            get_in = price - cost_sum
            sql1 = "INSERT INTO SALE (DATE,ID,NAME,MODEL,UNIT,COST,PRICE,NUM,COST_SUM,SUM,GET) \
                                      VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
                   (date, result1[0], result1[1], result1[3], result1[4], result1[5], result1[6], value[4], cost_sum,
                    price, get_in)
            status = self.exec_sqlite('w', sql1)
            if status == 1:
                self.put_sale_data()

    def clean_sale(self):
        m = self.table_sell_out.rowCount()
        for n in range(m):
            self.table_sell_out.removeRow(0)
        conn = sqlite3.connect(sql_file)
        c = conn.cursor()
        c.execute('''DROP TABLE SELL''')
        conn.commit()
        c.execute('''CREATE TABLE SELL
                                    (ID             INT     NOT NULL,
                                       NAME           TEXT    DEFAULT NULL,
                                       MODEL          TEXT    DEFAULT NULL,
                                       UNIT           TEXT    DEFAULT NULL,
                                       NUM            INT     DEFAULT NULL,
                                       PRICE          INT     DEFAULT NULL,
                                       SUM            INT     DEFAULT NULL,
                                       PRIMARY KEY (`ID`));''')
        conn.commit()
        conn.close()

    def put_setting_data(self):
        sql = "SELECT *  from INFO"
        info = self.exec_sqlite('r', sql)[0]
        name = info[0]
        locale = info[1]
        number = info[2]
        self.text_settings_1.setText(name)
        self.text_settings_2.setText(locale)
        self.text_settings_3.setText(number)

    def settings(self):
        name = self.text_settings_1.text()
        locale = self.text_settings_2.text()
        number = self.text_settings_3.text()
        sql = "UPDATE INFO set NAME = '%s',LOCALE = '%s',NUMBER = '%s'" % (name, locale, number)
        status = self.exec_sqlite('w', sql)
        if status == 1:
            QMessageBox.information(self, '提示！', '更新成功！')

    def excel_template(self):
        path, ok = QFileDialog.getSaveFileName(self, '保存文件', desktop_path + '/模板文件.xlsx', "Excel 文件(*.xlsx)")
        if ok == '':
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        title = ['商品id', '名称', '余货量', '型号', '单位', '成本价', '单价']
        ws.append(title)
        ws.cell(2, 1).value = 1
        ws.cell(3, 1).value = 2
        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.warning(self, '警告！', '导出失败，请检查文件是否被占用！')
        else:
            QMessageBox.information(self, '提示！', '导出成功！')
        finally:
            wb.close()

    def read_excel(self):
        path, ok = QFileDialog.getOpenFileName(self, '读取文件', desktop_path, "Excel 文件(*.xlsx)")
        if ok == '':
            return
        workbook = load_workbook(path)
        ws = workbook.active
        row = ws.max_row
        cell = []
        k = self.table_new_goods.rowCount()
        for i in range(k):
            self.table_new_goods.removeRow(0)
        self.table_new_goods.setRowCount(row - 1)
        for n in range(row - 1):
            for m in range(7):
                cell.append(ws.cell(n + 2, m + 1).value)
                item = QTableWidgetItem()
                self.table_new_goods.setItem(n, m, item)
                item_add = self.table_new_goods.item(n, m)
                item_add.setText(str(cell[m]))
                item_add.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)
            cell = []
        self.tabWidget.setCurrentIndex(1)

    def save_excel(self):
        path, ok = QFileDialog.getSaveFileName(self, '保存文件', desktop_path + '/商品清单.xlsx', "Excel 文件(*.xlsx)")
        if ok == '':
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        title = ['商品id', '名称', '余货量', '型号', '单位', '成本价', '单价']
        ws.append(title)
        sql = "SELECT *  from COMPANY"
        results = self.exec_sqlite('r', sql)
        n = 1
        for row in results:
            n = n + 1
            for i in range(len(row)):
                ws.cell(n, i + 1).value = row[i]
        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.warning(self, '警告！', '导出失败，请检查文件是否被占用！')
        else:
            QMessageBox.information(self, '提示！', '导出成功！')
        finally:
            wb.close()

    def save_sellout_excel_by_time(self):
        text, ok = QInputDialog.getText(self, '导出记录时间', '例如：导出某一年：2022   导出某一月：2022.4   导出某一天：2022.4.1',
                                        QLineEdit.Normal)
        if not ok:
            return
        date = text.split(sep='.', maxsplit=3)
        if len(date) == 0:
            QMessageBox.warning(self, '警告！', '请输入导出记录的时间！')
            return
        elif len(date) == 1:
            d = date[0]
        elif len(date) == 2:
            d = '%s/%s' % (date[0], date[1])
        elif len(date) == 3:
            d = '%s/%s/%s' % (date[0], date[1], date[2])
        else:
            QMessageBox.warning(self, '警告！', '请输入正确的时间！')
            return
        path, ok = QFileDialog.getSaveFileName(self, '保存文件', desktop_path + '/' + text + '-销售记录.xlsx',
                                               "Excel 文件(*.xlsx)")
        if ok == '':
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        title = ['日期', '商品id', '名称', '型号', '单位', '成本价', '单价', '售出数量', '总成本价', '总销售额', '利润']
        ws.append(title)
        sql = "SELECT * FROM SALE WHERE date like '%" + d + "%'"
        results = self.exec_sqlite('r', sql)
        n = 1
        for row in results:
            n = n + 1
            for i in range(len(row)):
                ws.cell(n, i + 1).value = row[i]
        rows = ws.max_row
        ws.cell(rows + 1, 8).value = '合计'
        ws.cell(rows + 1, 9).value = '=SUM(I2:I' + str(rows) + ')'
        ws.cell(rows + 1, 10).value = '=SUM(J2:J' + str(rows) + ')'
        ws.cell(rows + 1, 11).value = '=SUM(K2:K' + str(rows) + ')'
        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.warning(self, '警告！', '导出失败，请检查文件是否被占用！')
        else:
            QMessageBox.information(self, '提示！', '导出成功！')
        finally:
            wb.close()

    def save_sellout_excel(self):
        path, ok = QFileDialog.getSaveFileName(self, '保存文件', desktop_path + '/销售记录.xlsx', "Excel 文件(*.xlsx)")
        if ok == '':
            return
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        title = ['日期', '商品id', '名称', '型号', '单位', '成本价', '单价', '售出数量', '总成本价', '总销售额', '利润']
        ws.append(title)
        sql = "SELECT *  from SALE"
        results = self.exec_sqlite('r', sql)
        n = 1
        for row in results:
            n = n + 1
            for i in range(len(row)):
                ws.cell(n, i + 1).value = row[i]
        rows = ws.max_row
        ws.cell(rows + 1, 8).value = '合计'
        ws.cell(rows + 1, 9).value = '=SUM(I2:I' + str(rows) + ')'
        ws.cell(rows + 1, 10).value = '=SUM(J2:J' + str(rows) + ')'
        ws.cell(rows + 1, 11).value = '=SUM(K2:K' + str(rows) + ')'
        try:
            wb.save(path)
        except PermissionError:
            QMessageBox.warning(self, '警告！', '导出失败，请检查文件是否被占用！')
        else:
            QMessageBox.information(self, '提示！', '导出成功！')
        finally:
            wb.close()

    def get_update(self):
        url = 'https://api.azzb.workers.dev/https://api.github.com/repos/771073216/py/releases/latest'
        get_url = get(url)
        remote_version = str(re.findall('"tag_name":".....', get_url.text)).split('"')[3]
        if version == remote_version:
            QMessageBox.information(self, '提示！', '已是最新版本！')
            return
        self.up.show()

        # noinspection PyUnresolvedReferences
        self.download_update.download_process_signal.connect(self.set_progressbar_value)
        self.download_update.start()

    # 设置进度条
    def set_progressbar_value(self, value, size, current_size):
        self.up.progressBar.setValue(value)
        self.up.label.setText(
            str(format(current_size / 1000000, '.2f')) + ' MB /' + str(format(size / 1000000, '.2f')) + ' MB')
        if value >= 100:
            self.up.progressBar.setValue(100)
            QCoreApplication.quit()
            os.system(update_file)
            os.remove(update_file)
            os.system('"C:/Program Files/qt6shangpinguanli/start.exe"')

    def update_calculate_data(self, a):
        ok = bool()
        num, ok = QInputDialog.getInt(self, '输入框', '输入数字', ok)
        if ok is False:
            return
        elif ok is True and num <= 0:
            QMessageBox.information(self, '提示！', '请输入一个大于 0 的数！')
        else:
            sql = "UPDATE CAL set " + a + " = '%s'" % num
            self.exec_sqlite('w', sql)
            self.put_calculate_data()
            self.calculate()

    def put_calculate_data(self):
        sql = "SELECT *  from CAL"
        result = self.exec_sqlite('r', sql)[0]
        var_a = []
        var_b = []
        for i in range(7):
            var_a.append(str(result[i]))
            exec("self.text_a%s.setText(var_a[%s])" % (i + 1, i))
        for i in range(12):
            var_b.append(str(result[i + 7]))
            exec("self.text_b%s.setText(var_b[%s])" % (i + 1, i))

    @staticmethod
    def float_mul(a, b):
        c = float(a) * float(b)
        if c != int(c):
            c = int(c + 1)
        return str(int(c))

    def calculate(self):
        input_1 = self.text_cal_1.text()
        input_2 = self.text_cal_2.text()
        if input_1 == '' or input_2 == '' or not input_1 or not input_2:
            return

        self.text_cal_3.setText(str(format((float(input_1) + float(input_2)) * 2, '.2f')))
        self.text_cal_4.setText(str(format(float(input_1) * float(input_2), '.2f')))

        self.number_a1.setText(str(format(float(self.text_cal_3.text()) * 2 - 2, '.2f')))
        self.price_a1.setText(self.float_mul(self.number_a1.text(), self.text_a1.text()))
        self.number_a2.setText(self.text_cal_3.text())
        self.price_a2.setText(self.float_mul(self.number_a2.text(), self.text_a2.text()))
        self.number_a3.setText(self.text_cal_3.text())
        self.price_a3.setText(self.float_mul(self.number_a3.text(), self.text_a3.text()))
        self.number_a4.setText(str(format(float(self.text_cal_3.text()) - 2, '.2f')))
        self.price_a4.setText(self.float_mul(self.number_a4.text(), self.text_a4.text()))
        self.number_a5.setText(str(format(float(self.text_cal_3.text()) - 4, '.2f')))
        self.price_a5.setText(self.float_mul(self.number_a5.text(), self.text_a5.text()))
        self.price_a6.setText(self.float_mul(self.number_a6.text(), self.text_a6.text()))
        self.price_a7.setText(self.float_mul(self.number_a7.text(), self.text_a7.text()))
        self.text_sum_a.setText(str(int(self.price_a1.text()) + int(self.price_a2.text()) + int(self.price_a3.text())
                                    + int(self.price_a4.text()) + int(self.price_a5.text()) + int(self.price_a6.text())
                                    + int(self.price_a7.text())))

        self.number_b1.setText(self.text_cal_3.text())
        self.price_b1.setText(self.float_mul(self.number_b1.text(), self.text_b1.text()))
        self.number_b2.setText(str(format(float(self.text_cal_3.text()) * 2 - 2, '.2f')))
        self.price_b2.setText(self.float_mul(self.number_b2.text(), self.text_b2.text()))
        self.number_b3.setText(str(format(float(self.number_b2.text()) * 2, '.2f')))
        self.price_b3.setText(self.float_mul(self.number_b3.text(), self.text_b3.text()))
        self.number_b4.setText(self.text_cal_3.text())
        self.price_b4.setText(self.float_mul(self.number_b4.text(), self.text_b4.text()))
        self.number_b5.setText(str(format(float(self.text_cal_3.text()) - 2, '.2f')))
        self.price_b5.setText(self.float_mul(self.number_b5.text(), self.text_b5.text()))
        self.number_b6.setText(str(format(float(self.number_b4.text()) + float(self.number_b5.text()) - 4, '.2f')))
        self.price_b6.setText(self.float_mul(self.number_b6.text(), self.text_b6.text()))
        self.number_b7.setText(str(format(float(self.text_cal_3.text()) - 4, '.2f')))
        self.price_b7.setText(self.float_mul(self.number_b7.text(), self.text_b7.text()))
        self.number_b8.setText(str(int(float(self.number_b2.text()) / 5) + 1))
        self.price_b8.setText(self.float_mul(self.number_b8.text(), self.text_b8.text()))
        self.price_b9.setText(self.float_mul(self.number_b9.text(), self.text_b9.text()))
        self.price_b10.setText(self.float_mul(self.number_b10.text(), self.text_b10.text()))
        self.price_b11.setText(self.float_mul(self.number_b11.text(), self.text_b11.text()))
        self.price_b12.setText(self.float_mul(self.number_b12.text(), self.text_b12.text()))
        self.text_sum_b.setText(str(int(self.price_b1.text()) + int(self.price_b2.text()) + int(self.price_b3.text())
                                    + int(self.price_b4.text()) + int(self.price_b5.text()) + int(self.price_b6.text())
                                    + int(self.price_b7.text()) + int(self.price_b8.text()) + int(self.price_b9.text())
                                    + int(self.price_b10.text()) + int(self.price_b11.text())
                                    + int(self.price_b12.text())))

    def put_printer_data(self):
        printers = []
        for i in EnumPrinters(2):
            printers.append(i[2])
        self.comboBox.addItems(printers)
        sql = "SELECT PRINTER  from INFO"
        result = self.exec_sqlite('r', sql)[0]
        if not result[0] or result[0] == '':
            sql = "UPDATE INFO set PRINTER = '%s'" % GetDefaultPrinter()
            self.exec_sqlite('w', sql)
            sql = "SELECT PRINTER  from INFO"
            result = self.exec_sqlite('r', sql)[0]
        self.comboBox.setCurrentIndex(printers.index(result[0]))
        self.label_11.setText(result[0])

    def update_printer(self):
        sql = "UPDATE INFO set PRINTER = '%s'" % self.comboBox.currentText()
        status = self.exec_sqlite('w', sql)
        if status == 1:
            self.label_11.setText(self.comboBox.currentText())
            QMessageBox.information(self, '提示！', '更新成功！')


class DownloadUpdate(QThread):
    download_process_signal = Signal(int, float, int)

    def __init__(self):
        super(DownloadUpdate, self).__init__()

    def run(self):
        url = 'https://api.azzb.workers.dev/https://github.com/771073216/py/releases/latest/download/mysetup.exe'
        r = get(url, stream=True)
        data_size = int(r.headers['Content-Length'])
        i = 0
        f = open(update_file, 'wb')
        for chunk in r.iter_content(chunk_size=2048):
            if chunk:
                i = i + 1
                size = int(i * 2048 * 100 / data_size)
                # noinspection PyUnresolvedReferences
                self.download_process_signal.emit(size, data_size, i * 2048)
                f.write(chunk)
        f.close()


class Update(QDialog, Ui_Dialog):
    def __init__(self):
        super(Update, self).__init__()
        self.setupUi(self)


if __name__ == "__main__":
    import sys

    QGuiApplication.setHighDpiScaleFactorRoundingPolicy(Qt.HighDpiScaleFactorRoundingPolicy.Floor)
    app = QApplication(sys.argv)
    w = FirstWindow()
    w.show()
    sys.exit(app.exec())
