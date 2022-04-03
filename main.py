import os
import shutil
import openpyxl
import sqlite3
import datetime
import threading
from tkinter import ttk
import tkinter as tk
from tkinter import *
import tkinter.messagebox as messagebox
from tkinter import filedialog
from tkinter import simpledialog


def backup():
    sql_dir = os.path.expanduser("~/sql/")
    backup_dir = os.path.expanduser("~/sql/backup")
    is_backup_dir = os.path.exists(backup_dir)
    if not is_backup_dir:
        os.makedirs(backup_dir)
    i = datetime.datetime.now()
    date1 = "%s-%s-%s" % (i.year, i.month, i.day)
    pdir = os.path.expanduser("~/sql/backup")
    ppdir = pdir + '/' + date1
    isdir = os.path.exists(ppdir)
    if not isdir:
        os.makedirs(ppdir)
        shutil.copyfile(sql_dir + 'list.db', ppdir + '/' + 'list.db')


def check_sql():
    pdir = os.path.expanduser("~/sql/")
    isdir = os.path.exists(pdir)

    if not isdir:
        os.makedirs(pdir)
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        c.execute('''CREATE TABLE COMPANY
            (ID            INT     NOT NULL,
               NAME           TEXT    DEFAULT NULL,
               REMINDER       INT     DEFAULT NULL,
               MODEL          TEXT    DEFAULT NULL,
               UNIT           TEXT    DEFAULT NULL,
               COST           INT     DEFAULT NULL,
               PRICE          INT     DEFAULT NULL,
               PRIMARY KEY (`ID`));''')
        conn.commit()
        c.execute('''CREATE TABLE SALE
                (DATE             TEXT    NOT NULL,
                   ID             INT     NOT NULL,
                   NAME           TEXT    DEFAULT NULL,
                   MODEL          TEXT    DEFAULT NULL,
                   UNIT           TEXT    DEFAULT NULL,
                   COST           INT     DEFAULT NULL,
                   PRICE          INT     DEFAULT NULL,
                   NUM            INT     DEFAULT NULL,
                   COST_SUM       INT     DEFAULT NULL,
                   SUM            INT     DEFAULT NULL,
                   GET            INT     DEFAULT NULL);''')
        conn.commit()
        c.execute('''CREATE TABLE SELL
                        (ID             INT     NOT NULL,
                           NAME           TEXT    DEFAULT NULL,
                           MODEL          TEXT    DEFAULT NULL,
                           UNIT           TEXT    DEFAULT NULL,
                           NUM            INT     DEFAULT NULL,
                           PRICE          INT     DEFAULT NULL,
                           SUM            INT     DEFAULT NULL,
                           PRIMARY KEY (`ID`));''')
        conn.commit()
        conn.close()


def delete_sellout():
    conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
    c = conn.cursor()
    c.execute('''DROP TABLE SELL''')
    conn.commit()
    c.execute('''CREATE TABLE SELL
                            (ID             INT     NOT NULL,
                               NAME           TEXT    DEFAULT NULL,
                               MODEL          TEXT    DEFAULT NULL,
                               UNIT           TEXT    DEFAULT NULL,
                               NUM            INT     DEFAULT NULL,
                               PRICE          INT     DEFAULT NULL,
                               SUM            INT     DEFAULT NULL,
                               PRIMARY KEY (`ID`));''')
    conn.commit()
    conn.close()


def delete_sell():
    res = messagebox.askyesnocancel('警告！', '是否删除？')
    if res:
        res1 = messagebox.askyesnocancel('再次警告！', '是否删除？')
        if res1:
            conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
            c = conn.cursor()
            c.execute('''DROP TABLE SALE''')
            conn.commit()
            c.execute('''CREATE TABLE SALE
                (DATE             TEXT    NOT NULL,
                   ID             INT     NOT NULL,
                   NAME           TEXT    DEFAULT NULL,
                   MODEL          TEXT    DEFAULT NULL,
                   UNIT           TEXT    DEFAULT NULL,
                   COST           INT     DEFAULT NULL,
                   PRICE          INT     DEFAULT NULL,
                   NUM            INT     DEFAULT NULL,
                   COST_SUM       INT     DEFAULT NULL,
                   SUM            INT     DEFAULT NULL,
                   GET            INT     DEFAULT NULL);''')
            conn.commit()
            conn.close()
            messagebox.showinfo('提示！', '删除成功！')


def excel_template():
    td = threading.Thread(target=open_window)
    td.Daemon = True
    td.start()
    path = filedialog.asksaveasfilename(title=u'保存文件', initialfile='模板文件.xlsx',
                                        filetypes=(("Excel files", "*.xlsx"),))
    if not path:
        return
    str1 = path.split(sep=".", maxsplit=-1)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet 1"
    title = ['商品id', '名称', '余货量', '型号', '单位', '成本价', '单价']
    ws.append(title)
    ws.cell(2, 1).value = 1
    ws.cell(3, 1).value = 2
    try:
        wb.save(str1[0] + '.xlsx')
    except PermissionError:
        messagebox.showinfo('警告！', '导出失败，请检查文件是否被占用！')
    else:
        messagebox.showinfo('提示！', '导出成功！')


# 商品信息操作界面
def open_window():
    root = tk.Tk()
    root.withdraw()


class AdminPage:
    def __init__(self, parent_window):
        self.result = None
        self.row_info = None
        self.row = None
        parent_window.update()
        parent_window.destroy()  # 自动销毁上一个界面
        check_sql()
        backup()
        self.window = Tk()  # 初始框的声明
        self.window.title('管理面板')
        self.window.geometry("800x570+300+30")  # 初始窗口在屏幕中的位置
        self.frame_left_top = tk.Frame(width=300, height=240)  # 指定框架，在窗口上可以显示，这里指定四个框架
        self.frame_right_top = tk.Frame(width=350, height=220)
        self.frame_center = tk.Frame(width=800, height=200)
        self.frame_bottom = tk.Frame(width=800, height=50)

        # 菜单栏
        menubar = Menu(self.window, tearoff=0)
        menu1 = Menu(self.window, tearoff=0)
        menu2 = Menu(self.window, tearoff=0)
        menu3 = Menu(self.window, tearoff=0)
        menu4 = Menu(self.window, tearoff=0)
        menu5 = Menu(self.window, tearoff=0)
        menu6 = Menu(self.window, tearoff=0)

        menubar.add_cascade(label="菜单", menu=menu1)
        menu1.add_command(label='生成导入模板', command=excel_template, font=('微软雅黑', 15))

        menubar.add_cascade(label="导入", menu=menu2)
        menu2.add_command(label='从Excel文件导入商品信息', command=self.read_excel, font=('微软雅黑', 15))

        menubar.add_cascade(label="导出", menu=menu3)
        menu3.add_command(label='导出商品信息到Excel', command=self.save_excel, font=('微软雅黑', 15))
        menu3.add_command(label='导出全部销售记录到Excel', command=self.save_sellout_excel, font=('微软雅黑', 15))
        menu3.add_command(label='选择时间导出销售记录到Excel', command=self.save_sellout_excel_by_time, font=('微软雅黑', 15))
        menubar.add_cascade(label="               ", menu=menu4)
        menubar.add_cascade(label="               ", menu=menu5)
        menubar.add_cascade(label="高危菜单", menu=menu6)
        menu6.add_command(label='删除所有商品信息', command=self.delete_goods, font=('微软雅黑', 15))
        menu6.add_command(label='删除所有销售信息', command=delete_sell, font=('微软雅黑', 15))

        self.window['menu'] = menubar

        # 定义下方中心列表区域
        self.columns = ("商品id", "名称", "余货量", "型号", "单位", "成本价", "单价")
        self.tree = ttk.Treeview(self.frame_center, show="headings", height=8, columns=self.columns)
        # 添加竖直滚动条
        self.vbar = ttk.Scrollbar(self.frame_center, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)

        # 定义id1为修改id时的暂存变量，这个是为了更新信息而设计的
        self.id1 = 0
        # 表格的标题
        self.tree.column("商品id", width=100, anchor='center')
        self.tree.column("名称", width=150, anchor='center')
        self.tree.column("余货量", width=100, anchor='center')
        self.tree.column("型号", width=100, anchor='center')
        self.tree.column("单位", width=100, anchor='center')
        self.tree.column("成本价", width=100, anchor='center')
        self.tree.column("单价", width=100, anchor='center')

        # grid方法将tree和vbar进行布局
        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)

        self.date = []
        self.num = []
        self.sum = []

        # 定义几个数组，为中间的那个大表格做一些准备

        self.id = []
        self.name = []
        self.reminder = []
        self.model = []
        self.unit = []
        self.cost = []
        self.price = []
        # 打开数据库连接
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from COMPANY")
        results = cursor.fetchall()
        print(results)
        # 向表格中插入数据
        for row in results:
            self.id.append(row[0])
            self.name.append(row[1])
            self.reminder.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])

        for i in range(min(len(self.id), len(self.name), len(self.reminder),
                           len(self.model), len(self.unit), len(self.cost), len(self.price))):  # 写入数据
            self.tree.insert('', i, values=(self.id[i], self.name[i], self.reminder[i],
                                            self.model[i], self.unit[i], self.cost[i], self.price[i]))

        # 绑定函数，使表头可排序
        for col in self.columns:
            self.tree.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

        # 定义左上方区域
        self.left_top_frame = tk.Frame(self.frame_left_top)
        self.var_id = StringVar()  # 声明学号
        self.var_name = StringVar()  # 声明姓名
        self.var_reminder = StringVar()  # 声明性别
        self.var_model = StringVar()  # 声明年龄
        self.var_unit = StringVar()  # 声明性别
        self.var_cost = StringVar()  # 声明性别
        self.var_price = StringVar()  # 声明年龄
        # 商品id
        self.right_top_id_label = Label(self.frame_left_top, text="商品id： ", font=('微软雅黑', 15))
        self.right_top_id_entry = Entry(self.frame_left_top, textvariable=self.var_id, font=('微软雅黑', 15))
        self.right_top_id_label.grid(row=1, column=0)
        self.right_top_id_entry.grid(row=1, column=1)
        # 商品名称
        self.right_top_name_label = Label(self.frame_left_top, text="名称：", font=('微软雅黑', 15))
        self.right_top_name_entry = Entry(self.frame_left_top, textvariable=self.var_name, font=('微软雅黑', 15))
        self.right_top_name_label.grid(row=2, column=0)  # 位置设置
        self.right_top_name_entry.grid(row=2, column=1)
        # 余货量
        self.right_top_remainder_label = Label(self.frame_left_top, text="余货量：", font=('微软雅黑', 15))
        self.right_top_remainder_entry = Entry(self.frame_left_top, textvariable=self.var_reminder, font=('微软雅黑', 15))
        self.right_top_remainder_label.grid(row=3, column=0)  # 位置设置
        self.right_top_remainder_entry.grid(row=3, column=1)
        # 型号
        self.right_top_model_label = Label(self.frame_left_top, text="型号：", font=('微软雅黑', 15))
        self.right_top_model_entry = Entry(self.frame_left_top, textvariable=self.var_model, font=('微软雅黑', 15))
        self.right_top_model_label.grid(row=4, column=0)  # 位置设置
        self.right_top_model_entry.grid(row=4, column=1)

        self.right_top_unit_label = Label(self.frame_left_top, text="单位：", font=('微软雅黑', 15))
        self.right_top_unit_entry = Entry(self.frame_left_top, textvariable=self.var_unit, font=('微软雅黑', 15))
        self.right_top_unit_label.grid(row=5, column=0)  # 位置设置
        self.right_top_unit_entry.grid(row=5, column=1)

        self.right_top_cost_label = Label(self.frame_left_top, text="成本价：", font=('微软雅黑', 15))
        self.right_top_cost_entry = Entry(self.frame_left_top, textvariable=self.var_cost, font=('微软雅黑', 15))
        self.right_top_cost_label.grid(row=6, column=0)  # 位置设置
        self.right_top_cost_entry.grid(row=6, column=1)

        self.right_top_price_label = Label(self.frame_left_top, text="单价：", font=('微软雅黑', 15))
        self.right_top_price_entry = Entry(self.frame_left_top, textvariable=self.var_price, font=('微软雅黑', 15))
        self.right_top_price_label.grid(row=7, column=0)  # 位置设置
        self.right_top_price_entry.grid(row=7, column=1)

        # 定义右上方区域
        # self.right_top_title = Label(self.frame_right_top, text="操作：", font=('微软雅黑', 15))
        self.tree.bind('<Button-1>', self.click)  # 左键获取位置(tree.bind可以绑定一系列的事件，可以搜索ttk相关参数查看)
        self.right_top_button1 = ttk.Button(self.frame_right_top, text='新建商品信息',
                                            width=20, command=self.new_row, padding=5)
        self.right_top_button2 = ttk.Button(self.frame_right_top, text='更新选中商品信息',
                                            width=20, command=self.update_row, padding=5)
        self.right_top_button3 = ttk.Button(self.frame_right_top, text='删除选中商品信息',
                                            width=20, command=self.del_row, padding=5)
        self.right_top_button3.bind_all('<Delete>', self.eventhandler)
        self.right_top_button4 = ttk.Button(self.frame_right_top, text='售出选中商品',
                                            width=20, command=self.edit_num, padding=5)
        self.right_top_button5 = ttk.Button(self.frame_right_top, text='查看账单',
                                            width=20, command=lambda: SellOut(self.window), padding=5)
        self.right_top_button6 = ttk.Button(self.frame_right_top, text='出货',
                                            width=20, command=lambda: NewSell(self.window), padding=5)

        # 定义下方区域，查询功能块
        self.search = StringVar()
        self.right_bottom_search_entry = Entry(self.frame_bottom, textvariable=self.search, font=('微软雅黑', 15))
        self.right_bottom_button = ttk.Button(self.frame_bottom, text='商品名称查询', width=20, command=self.put_data)
        self.right_bottom_button.grid(row=0, column=1, padx=20, pady=20)  # 位置设置
        self.right_bottom_search_entry.grid(row=0, column=0, padx=20, pady=20)

        self.right_bottom_button1 = ttk.Button(self.frame_bottom, text='清除查询结果', width=20, command=self.clean_search)
        self.right_bottom_button1.grid(row=0, column=2, padx=20, pady=20)

        # 右上角按钮的位置设置
        self.right_top_button1.grid(row=2, column=0, padx=5, pady=5)
        self.right_top_button2.grid(row=3, column=0, padx=5, pady=5)
        self.right_top_button3.grid(row=4, column=0, padx=5, pady=5)
        self.right_top_button3.bind_all('Delete', self.del_row)
        self.right_top_button4.grid(row=5, column=0, padx=5, pady=5)
        self.right_top_button5.grid(row=3, column=1, padx=5, pady=5)
        self.right_top_button6.grid(row=2, column=1, padx=5, pady=5)

        # 整体区域定位，利用了Frame和grid进行布局
        self.frame_left_top.grid(row=0, column=0, padx=2, pady=5)
        self.frame_right_top.grid(row=0, column=1, padx=30, pady=30)
        self.frame_center.grid(row=1, column=0, columnspan=2, padx=5, pady=5)
        self.frame_bottom.grid(row=2, column=0, columnspan=2)

        # 设置固定组件，(0)就是将组件进行了固定
        self.frame_left_top.grid_propagate(False)
        self.frame_right_top.grid_propagate(False)
        self.frame_center.grid_propagate(False)
        self.frame_bottom.grid_propagate(False)

        self.frame_left_top.tkraise()  # 开始显示主菜单，tkraise()提高z轴的顺序（不太懂）
        self.frame_right_top.tkraise()  # 开始显示主菜单
        self.frame_center.tkraise()  # 开始显示主菜单
        self.frame_bottom.tkraise()  # 开始显示主菜单

        self.window.protocol("WM_DELETE_WINDOW", self.back)  # 捕捉右上角关闭点击，执行back方法
        self.window.mainloop()  # 进入消息循环

    def clean_search(self):
        self.search.set('')
        self.put_data()

    # 将查到的信息放到中间的表格中
    def put_data(self):
        self.del_button()  # 先将表格内的内容全部清空
        # 打开数据库连接，准备查找指定的信息
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        d = self.search.get()
        e = conn.cursor().execute("SELECT * FROM COMPANY WHERE name like '%" + d + "%'")
        results = e.fetchall()
        self.id = []
        self.name = []
        self.reminder = []
        self.model = []
        self.unit = []
        self.cost = []
        self.price = []
        # 向表格中插入数据
        for row in results:
            self.id.append(row[0])
            self.name.append(row[1])
            self.reminder.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])

        for i in range(min(len(self.id), len(self.name), len(self.reminder),
                           len(self.model), len(self.unit), len(self.cost), len(self.price))):  # 写入数据
            self.tree.insert('', i, values=(self.id[i], self.name[i], self.reminder[i],
                                            self.model[i], self.unit[i], self.cost[i], self.price[i]))

        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

    # 清空表格中的所有信息
    def del_button(self):
        x = self.tree.get_children()
        for item in x:
            self.tree.delete(item)

    # 在表格上的点击事件，这里是作用就是一点击表格就可以将信息直接写到左上角的框框中
    def click(self, event):
        if not self.tree.identify_row(event.y):
            return
        self.row = self.tree.identify_row(event.y)  # 行
        self.row_info = self.tree.item(self.row, "values")
        self.var_id.set(self.row_info[0])
        self.id1 = self.var_id.get()
        self.var_name.set(self.row_info[1])
        self.var_reminder.set(self.row_info[2])
        self.var_model.set(self.row_info[3])
        self.var_unit.set(self.row_info[4])
        self.var_cost.set(self.row_info[5])
        self.var_price.set(self.row_info[6])
        self.right_top_id_entry = Entry(self.frame_left_top, state='disabled', textvariable=self.var_id,
                                        font=('微软雅黑', 15))

    # 点击中间的表格的表头，可以将那一列进行排序
    def tree_sort_column(self, tv, col, reverse):  # Treeview、列名、排列方式
        m = [(tv.set(k, col), k) for k in tv.get_children('')]
        m.sort(reverse=reverse)  # 排序方式
        for index, (val, k) in enumerate(m):  # 根据排序后索引移动
            tv.move(k, '', index)
        tv.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(tv, _col, not reverse))

    def new_row(self):
        if self.var_id.get() == '' or self.var_name.get() == '' or self.var_reminder.get() == '' \
                or self.var_model.get() == '' or self.var_unit.get() == '' or self.var_cost.get() == '' \
                or self.var_price.get() == '':
            messagebox.showinfo('警告！', '请填写商品信息')
        else:
            if str(self.var_id.get()) in self.id:
                messagebox.showinfo('警告！', '该商品已存在！')
            else:
                # 打开数据库连接
                conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
                c = conn.cursor()
                sql = "INSERT INTO COMPANY (ID,NAME,REMINDER,MODEL,UNIT,COST,PRICE) \
                           VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
                      (self.var_id.get(), self.var_name.get(), self.var_reminder.get(),
                       self.var_model.get(), self.var_unit.get(), self.var_cost.get(), self.var_price.get())
                try:
                    sqlerror = "0"
                    c.execute(sql)  # 执行sql语句
                    conn.commit()  # 提交到数据库执行
                except sqlite3.IntegrityError:
                    sqlerror = "1"
                    messagebox.showinfo('警告！', '该商品id已存在！')
                except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                    sqlerror = "1"
                    conn.rollback()
                    messagebox.showinfo('警告！', '插入失败，数据库写入操作失败！')
                conn.close()
                if sqlerror == "0":
                    self.id.append(self.var_id.get())
                    self.name.append(self.var_name.get())
                    self.reminder.append(self.var_reminder.get())
                    self.model.append(self.var_model.get())
                    self.unit.append(self.var_unit.get())
                    self.cost.append(self.var_cost.get())
                    self.price.append(self.var_price.get())
                    self.tree.insert('', len(self.id) - 1, values=(
                        self.id[len(self.id) - 1], self.name[len(self.id) - 1], self.reminder[len(self.id) - 1],
                        self.model[len(self.id) - 1], self.unit[len(self.id) - 1], self.cost[len(self.id) - 1],
                        self.price[len(self.id) - 1]))
                    self.tree.update()
                    messagebox.showinfo('提示！', '插入成功！')

    def update_row(self):
        if self.var_id.get() == '':
            messagebox.showinfo('警告！', '请点击要更新的项目！')
        else:
            if self.var_id.get() == '' or self.var_name.get() == '' or self.var_reminder.get() == '' \
                    or self.var_model.get() == '' or self.var_unit.get() == '' or self.var_cost.get() == '' \
                    or self.var_price.get() == '':
                messagebox.showinfo('警告！', '请填写商品信息')
            else:
                res = messagebox.askyesnocancel('警告！', '是否更新所填数据？')
                if res:
                    # 打开数据库连接
                    conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
                    c = conn.cursor()
                    sql = "UPDATE COMPANY set ID = '%s', NAME = '%s', REMINDER = '%s', MODEL='%s' ," \
                          " UNIT = '%s', COST='%s', PRICE='%s' where ID='%s'" % \
                          (self.var_id.get(), self.var_name.get(), self.var_reminder.get(),
                           self.var_model.get(), self.var_unit.get(), self.var_cost.get(), self.var_price.get(),
                           self.id1)
                    try:
                        sqlerror = "0"
                        c.execute(sql)
                        conn.commit()
                    except sqlite3.IntegrityError:
                        sqlerror = "1"
                        messagebox.showinfo('警告！', '该商品id已存在！')
                    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                        sqlerror = "1"
                        conn.rollback()
                        messagebox.showinfo('警告！', '更新失败，数据库写入操作失败！')
                    conn.close()
                    if sqlerror == "0":
                        self.tree.item(self.tree.selection()[0], values=(
                            self.var_id.get(), self.var_name.get(), self.var_reminder.get(),
                            self.var_model.get(), self.var_unit.get(), self.var_cost.get(), self.var_price.get()))
                    messagebox.showinfo('提示！', '更新成功！')

    def eventhandler(self, event):
        if event.keysym == 'Delete':
            self.del_row()

    # 删除行
    def del_row(self):
        if self.var_id.get() == '':
            messagebox.showinfo('警告！', '请点击要删除的项目！')
        else:
            res = messagebox.askyesnocancel('警告！', '是否删除所选数据？')
            if res:
                # 打开数据库连接
                conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
                c = conn.cursor()
                sql = "DELETE from COMPANY where ID='%s'" % (self.row_info[0])
                try:
                    c.execute(sql)
                    conn.commit()
                except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                    conn.rollback()
                    messagebox.showinfo('警告！', '删除失败，数据库删除操作失败！')
                else:
                    conn.close()
                    self.tree.delete(self.tree.selection()[0])  # 删除所选行
                    self.clean_info()
                    messagebox.showinfo('提示！', '删除成功！')

    def clean_info(self):
        self.var_id.set('')
        self.var_name.set('')
        self.var_reminder.set('')
        self.var_model.set('')
        self.var_unit.set('')
        self.var_cost.set('')
        self.var_price.set('')

    def sellout(self):
        i = datetime.datetime.now()
        date1 = "%s/%s/%s %s:%s:%s" % (i.year, i.month, i.day, i.hour, i.minute, i.second)
        uuu = int(self.row_info[6])
        price = uuu * self.result
        print(self.row_info[5])
        print(self.result)
        cost_sum = int(self.row_info[5]) * int(self.result)
        print(cost_sum)
        get = int(price) - int(cost_sum)
        print(get)
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        sql = "INSERT INTO SALE (DATE,ID,NAME,MODEL,UNIT,COST,PRICE,NUM,COST_SUM,SUM,GET) \
                                   VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
              (date1, self.row_info[0], self.row_info[1],
               self.row_info[3], self.row_info[4], self.row_info[5], self.row_info[6], self.result, cost_sum, price,
               get)
        try:
            c.execute(sql)  # 执行sql语句
            conn.commit()  # 提交到数据库执行
        except (sqlite3.IntegrityError, sqlite3.ProgrammingError, sqlite3.OperationalError):
            conn.rollback()
            conn.close()
            messagebox.showinfo('警告！', '数据库写入操作失败！')
        conn.close()

    def edit_num(self):
        try:
            self.row_info = self.tree.item(self.row, "values")
        except TclError:
            messagebox.showinfo('警告！', '请选择要修改的项目')
            return FALSE
        td = threading.Thread(target=open_window)
        td.Daemon = True
        td.start()
        self.result = tk.simpledialog.askinteger(title=' ', prompt='售出数量：')
        if not self.result:
            return
        self.id1 = self.var_id.get()
        d = self.result
        m = self.row_info[2]
        m1 = int(m)
        if d <= 0:
            messagebox.showinfo('警告！', '请输入大于0的数')
            return FALSE
        k = m1 - d
        if k >= 0:
            # 打开数据库连接
            conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
            c = conn.cursor()
            sql = "UPDATE COMPANY set REMINDER = '%s' where ID='%s'" % \
                  (k, self.id1)
            try:
                c.execute(sql)
                conn.commit()
            except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                conn.rollback()
                conn.close()
                messagebox.showinfo('警告！', '更新失败，数据库写入操作失败！')
            else:
                conn.close()
                self.tree.item(self.tree.selection()[0], values=(
                    self.var_id.get(), self.var_name.get(), k,
                    self.var_model.get(), self.var_unit.get(), self.var_cost.get(), self.var_price.get()))
                k1 = str(k)
                self.var_reminder.set(k1)
                self.sellout()
                messagebox.showinfo('提示！', '更新成功！')
        else:
            messagebox.showinfo('警告！', '更新失败，余货量不能为负数！')

    def read_excel(self):
        # 打开文件
        td = threading.Thread(target=open_window)
        td.Daemon = True
        td.start()
        path = filedialog.askopenfilename()
        if not path:
            return
        workbook = openpyxl.load_workbook(path)
        m = workbook.active
        column = m.max_column
        row = m.max_row
        values = []
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        try:
            for i in range(2, (row + 1)):
                for j in range(1, (column + 1)):
                    values.append(m.cell(row=i, column=j).value)
                sql = "INSERT INTO COMPANY (ID,NAME,REMINDER,MODEL,UNIT,COST,PRICE) \
                                        VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
                      (values[0], values[1], values[2], values[3], values[4], values[5], values[6])
                c.execute(sql)  # 执行sql语句
                del values[:]
        except sqlite3.IntegrityError:
            messagebox.showinfo('警告！', '导入失败，请检查商品id是否重复！')
        else:
            conn.commit()
            self.clean_search()
            messagebox.showinfo('提示！', '导入成功！')

    def save_excel(self):
        # 打开文件
        td = threading.Thread(target=open_window)
        td.Daemon = True
        td.start()
        path = filedialog.asksaveasfilename(title=u'保存文件', initialfile='导出表单.xlsx',
                                            filetypes=(("Excel files", "*.xlsx"),))
        if not path:
            return
        str1 = path.split(sep=".xlsx", maxsplit=1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from COMPANY")
        results = cursor.fetchall()
        self.id1 = []
        self.name = []
        self.reminder = []
        self.model = []
        self.unit = []
        self.cost = []
        self.price = []
        # 向表格中插入数据
        for row in results:
            self.id1.append(row[0])
            self.name.append(row[1])
            self.reminder.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])
        n = 0
        title = ['商品id', '名称', '余货量', '型号', '单位', '成本价', '单价']
        ws.append(title)
        for var in [self.id1, self.name, self.reminder, self.model, self.unit, self.cost, self.price]:
            m = var[:]
            n = n + 1
            for i in range(len(m)):
                ws.cell(i + 2, n).value = m[i]
        try:
            wb.save(str1[0] + '.xlsx')
        except PermissionError:
            messagebox.showinfo('警告！', '导出失败，请检查文件是否被占用！')
        else:
            messagebox.showinfo('提示！', '导出成功！')

    def save_sellout_excel_by_time(self):
        td = threading.Thread(target=open_window)
        td.Daemon = True
        td.start()
        dd = tk.simpledialog.askstring(title='导出记录时间', prompt='例如：导出某一年：2022   导出某一月：2022.4   导出某一天：2022.4.1')
        if not dd:
            return
        date = dd.split(sep='.', maxsplit=3)
        print(date)
        if len(date) == 0:
            messagebox.showinfo('警告！', '请输入导出记录的时间！')
            return
        elif len(date) == 1:
            d = date[0]
        elif len(date) == 2:
            d = '%s/%s' % (date[0], date[1])
        elif len(date) == 3:
            d = '%s/%s/%s' % (date[0], date[1], date[2])
        else:
            messagebox.showinfo('警告！', '请输入正确的时间！')
            return
        path = filedialog.asksaveasfilename(title=u'保存文件', initialfile=dd + '-销售记录.xlsx',
                                            filetypes=(("Excel files", "*.xlsx"),))
        if not path:
            return
        str1 = path.split(sep=".xlsx", maxsplit=1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT * FROM SALE WHERE date like '%" + d + "%'")
        results = cursor.fetchall()
        # DATE,ID,NAME,MODEL,UNIT,COST,PRICE,NUM,COST_SUM,SUM,GET
        self.date = []
        self.id1 = []
        self.name = []
        self.model = []
        self.unit = []
        self.cost = []
        self.price = []
        self.num = []
        self.cost_sum = []
        self.sum = []
        self.get = []
        # 向表格中插入数据
        for row in results:
            self.date.append(row[0])
            self.id1.append(row[1])
            self.name.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])
            self.num.append(row[7])
            self.cost_sum.append(row[8])
            self.sum.append(row[9])
            self.get.append(row[10])
        n = 0
        title = ['日期', '商品id', '名称', '型号', '单位', '成本价', '单价', '售出数量', '总成本价', '总销售额', '利润']
        ws.append(title)
        for var in [self.date, self.id1, self.name, self.model, self.unit, self.cost, self.price, self.num,
                    self.cost_sum, self.sum, self.get]:
            m = var[:]
            n = n + 1
            for i in range(len(m)):
                ws.cell(i + 2, n).value = m[i]
        row1 = ws.max_row
        row = ws.max_row + 1
        print(row)
        ws.cell(row, 9).value = '=SUM(I2:I' + str(row1) + ')'
        ws.cell(row, 10).value = '=SUM(J2:J' + str(row1) + ')'
        ws.cell(row, 11).value = '=SUM(K2:K' + str(row1) + ')'
        try:
            wb.save(str1[0] + '.xlsx')
        except PermissionError:
            messagebox.showinfo('警告！', '导出失败，请检查文件是否被占用！')
        else:
            messagebox.showinfo('提示！', '导出成功！')

    def save_sellout_excel(self):
        # 打开文件
        td = threading.Thread(target=open_window)
        td.Daemon = True
        td.start()
        path = filedialog.asksaveasfilename(title=u'保存文件', initialfile='销售记录.xlsx',
                                            filetypes=(("Excel files", "*.xlsx"),))
        if not path:
            return
        str1 = path.split(sep=".", maxsplit=-1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from SALE")
        results = cursor.fetchall()
        self.date = []
        self.id1 = []
        self.name = []
        self.model = []
        self.unit = []
        self.cost = []
        self.price = []
        self.num = []
        self.cost_sum = []
        self.sum = []
        self.get = []
        # 向表格中插入数据
        for row in results:
            self.date.append(row[0])
            self.id1.append(row[1])
            self.name.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])
            self.num.append(row[7])
            self.cost_sum.append(row[8])
            self.sum.append(row[9])
            self.get.append(row[10])
        n = 0
        title = ['日期', '商品id', '名称', '型号', '单位', '成本价', '单价', '售出数量', '总成本价', '总销售额', '利润']
        ws.append(title)
        for var in [self.date, self.id1, self.name, self.model, self.unit, self.cost, self.price, self.num,
                    self.cost_sum, self.sum, self.get]:
            m = var[:]
            n = n + 1
            for i in range(len(m)):
                ws.cell(i + 2, n).value = m[i]
        row1 = ws.max_row
        row = ws.max_row + 1
        print(row)
        ws.cell(row, 9).value = '=SUM(I2:I' + str(row1) + ')'
        ws.cell(row, 10).value = '=SUM(J2:J' + str(row1) + ')'
        ws.cell(row, 11).value = '=SUM(K2:K' + str(row1) + ')'
        try:
            wb.save(str1[0] + '.xlsx')
        except PermissionError:
            messagebox.showinfo('警告！', '导出失败，请检查文件是否被占用！')
        else:
            messagebox.showinfo('提示！', '导出成功！')

    def delete_goods(self):
        res = messagebox.askyesnocancel('警告！', '是否删除？')
        if res:
            res1 = messagebox.askyesnocancel('再次警告！', '是否删除？')
            if res1:
                conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
                c = conn.cursor()
                c.execute('''DROP TABLE COMPANY''')
                conn.commit()
                c.execute('''CREATE TABLE COMPANY
                    (ID            INT     NOT NULL,
                       NAME           TEXT    DEFAULT NULL,
                       REMINDER       INT     DEFAULT NULL,
                       MODEL          TEXT    DEFAULT NULL,
                       UNIT           TEXT    DEFAULT NULL,
                       COST           INT     DEFAULT NULL,
                       PRICE          INT     DEFAULT NULL,
                       PRIMARY KEY (`ID`));''')
                conn.commit()
                conn.close()
                self.clean_search()
                self.clean_info()
                messagebox.showinfo('提示！', '删除成功！')

    def back(self):
        self.window.destroy()  # 进入管理员子菜单操作界面


class SellOut:
    def __init__(self, parent_window):
        self.row_info = None
        self.row = None
        parent_window.update()
        parent_window.destroy()  # 自动销毁上一个界面
        self.window = Tk()  # 初始框的声明
        self.window.title('管理面板')
        self.window.geometry("1000x550+300+30")  # 初始窗口在屏幕中的位置
        self.frame_center = tk.Frame(width=1200, height=200)
        # 定义下方中心列表区域
        self.columns = ('日期', '商品id', '名称', '型号', '单位', '成本价', '单价', '售出数量', '总成本价', '总销售额', '利润')
        self.tree = ttk.Treeview(self.frame_center, show="headings", height=8, columns=self.columns)
        # 添加竖直滚动条
        self.vbar = ttk.Scrollbar(self.frame_center, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)
        # 定义id1为修改id时的暂存变量，这个是为了更新信息而设计的
        self.id1 = 0
        # 表格的标题
        self.tree.column("日期", width=150, anchor='center')
        self.tree.column("商品id", width=50, anchor='center')
        self.tree.column("名称", width=100, anchor='center')
        self.tree.column("型号", width=100, anchor='center')
        self.tree.column("单位", width=50, anchor='center')
        self.tree.column("成本价", width=50, anchor='center')
        self.tree.column("单价", width=100, anchor='center')
        self.tree.column("售出数量", width=50, anchor='center')
        self.tree.column("总成本价", width=100, anchor='center')
        self.tree.column("总销售额", width=100, anchor='center')
        self.tree.column("利润", width=100, anchor='center')
        self.tree.bind('<Button-1>', self.click)
        # grid方法将tree和vbar进行布局
        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)
        # 定义几个数组，为中间的那个大表格做一些准备
        self.date = []
        self.id = []
        self.name = []
        self.model = []
        self.unit = []
        self.cost = []
        self.price = []
        self.num = []
        self.cost_sum = []
        self.sum = []
        self.get = []
        # 打开数据库连接
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from SALE")
        # DATE, ID, NAME, MODEL, UNIT, PRICE, NUM, SUM
        results = cursor.fetchall()
        for row in results:
            self.date.append(row[0])
            self.id.append(row[1])
            self.name.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])
            self.num.append(row[7])
            self.cost_sum.append(row[8])
            self.sum.append(row[9])
            self.get.append(row[10])
        conn.close()
        # 写入数据
        for i in range(min(len(self.date), len(self.id), len(self.name), len(self.model),
                           len(self.unit), len(self.cost), len(self.price), len(self.num), len(self.cost_sum),
                           len(self.sum), len(self.get))):
            self.tree.insert('', i, values=(self.date[i], self.id[i], self.name[i], self.model[i],
                                            self.unit[i], self.cost[i], self.price[i], self.num[i], self.cost_sum[i],
                                            self.sum[i], self.get[i]))
        # 绑定函数，使表头可排序
        for col in self.columns:
            self.tree.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))
        # 整体区域定位，利用了Frame和grid进行布局
        self.frame_bottom = tk.Frame(width=1000, height=300)
        self.search = StringVar()
        self.right_bottom_search_entry = Entry(self.frame_bottom, textvariable=self.search, font=('微软雅黑', 15))
        self.right_bottom_button = ttk.Button(self.frame_bottom, text='按名称查询',
                                              width=20, command=self.put_data, padding=5)
        self.right_bottom_button.grid(row=1, column=0, padx=0, pady=0)  # 位置设置
        self.right_bottom_search_entry.grid(row=0, column=0, padx=20, pady=20)
        self.right_bottom_title = Label(self.frame_bottom, text="按时间查询，例如：2022.4.3", font=('Verdana', 15))
        self.right_bottom_title.grid(row=1, column=2, pady=10)
        self.right_bottom_button1 = ttk.Button(self.frame_bottom, text='清除查询结果',
                                               width=20, command=self.clean_search, padding=5)
        self.right_bottom_button1.grid(row=0, column=1, padx=20, pady=20)
        self.right_bottom_button2 = ttk.Button(self.frame_bottom, text='按时间查询',
                                               width=20, command=self.put_data_time, padding=5)
        self.right_bottom_button2.grid(row=1, column=1, padx=0, pady=0)
        self.right_bottom_button4 = ttk.Button(self.frame_bottom, text='删除选中商品销售信息',
                                               width=20, command=self.del_row, padding=5)
        self.right_bottom_button4.grid(row=2, column=1, padx=20, pady=20)

        self.frame_bottom.grid(row=2, column=0, columnspan=2)

        self.frame_center.grid(row=1, column=0, columnspan=2, padx=4, pady=5)
        self.frame_center.grid_propagate(False)
        self.frame_bottom.grid_propagate(False)

        self.frame_center.tkraise()  # 开始显示主菜单
        self.frame_bottom.tkraise()
        self.window.protocol("WM_DELETE_WINDOW", self.back)  # 捕捉右上角关闭点击，执行back方法
        self.window.mainloop()  # 进入消息循环

    def clean_search(self):
        self.search.set('')
        self.put_data()

    def del_row(self):
        if not self.row_info:
            messagebox.showinfo('警告！', '请点击要删除的项目！')
        else:
            res = messagebox.askyesnocancel('警告！', '是否删除所选数据？')
            if res:
                # 打开数据库连接
                conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
                c = conn.cursor()
                sql = "DELETE from SALE where DATE='%s'" % (self.row_info[0])
                try:
                    c.execute(sql)
                    conn.commit()
                except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                    conn.rollback()
                    messagebox.showinfo('警告！', '删除失败，数据库删除操作失败！')
                else:
                    conn.close()
                    self.tree.delete(self.tree.selection()[0])  # 删除所选行
                    messagebox.showinfo('提示！', '删除成功！')

    # 将查到的信息放到中间的表格中
    def put_data(self):
        self.del_button()
        self.id = []
        self.name = []
        self.model = []
        self.unit = []
        self.price = []
        self.date = []
        self.num = []
        self.sum = []
        # 打开数据库连接，准备查找指定的信息
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        d = self.search.get()
        e = conn.cursor().execute("SELECT * FROM SALE WHERE name like '%" + d + "%'")
        results = e.fetchall()
        for row in results:
            self.date.append(row[0])
            self.id.append(row[1])
            self.name.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])
            self.num.append(row[7])
            self.cost_sum.append(row[8])
            self.sum.append(row[9])
            self.get.append(row[10])
        conn.close()
        # 写入数据
        for i in range(min(len(self.date), len(self.id), len(self.name), len(self.model),
                           len(self.unit), len(self.cost), len(self.price), len(self.num), len(self.cost_sum),
                           len(self.sum), len(self.get))):
            self.tree.insert('', i, values=(self.date[i], self.id[i], self.name[i], self.model[i],
                                            self.unit[i], self.cost[i], self.price[i], self.num[i], self.cost_sum[i],
                                            self.sum[i], self.get[i]))
        # 绑定函数，使表头可排序
        for col in self.columns:
            self.tree.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

    def put_data_time(self):
        self.del_button()
        self.id = []
        self.name = []
        self.model = []
        self.unit = []
        self.cost = []
        self.price = []
        self.num = []
        self.cost_sum = []
        self.sum = []
        self.get = []
        # 打开数据库连接，准备查找指定的信息
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        d = self.search.get().replace('.', '/')
        e = conn.cursor().execute("SELECT * FROM SALE WHERE date like '%" + d + "%'")
        results = e.fetchall()
        for row in results:
            self.date.append(row[0])
            self.id.append(row[1])
            self.name.append(row[2])
            self.model.append(row[3])
            self.unit.append(row[4])
            self.cost.append(row[5])
            self.price.append(row[6])
            self.num.append(row[7])
            self.cost_sum.append(row[8])
            self.sum.append(row[9])
            self.get.append(row[10])
        conn.close()
        # 写入数据
        for i in range(min(len(self.date), len(self.id), len(self.name), len(self.model),
                           len(self.unit), len(self.cost), len(self.price), len(self.num), len(self.cost_sum),
                           len(self.sum), len(self.get))):
            self.tree.insert('', i, values=(self.date[i], self.id[i], self.name[i], self.model[i],
                                            self.unit[i], self.cost[i], self.price[i], self.num[i], self.cost_sum[i],
                                            self.sum[i], self.get[i]))

        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

    def del_button(self):
        print('1')
        x = self.tree.get_children()
        print(x)
        for item in x:
            self.tree.delete(item)

    def tree_sort_column(self, tv, col, reverse):  # Treeview、列名、排列方式
        m = [(tv.set(k, col), k) for k in tv.get_children('')]
        m.sort(reverse=reverse)  # 排序方式
        for index, (val, k) in enumerate(m):  # 根据排序后索引移动
            tv.move(k, '', index)
        tv.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(tv, _col, not reverse))

    def click(self, event):
        if not self.tree.identify_row(event.y):
            return
        self.row = self.tree.identify_row(event.y)  # 行
        self.row_info = self.tree.item(self.row, "values")

    def back(self):
        AdminPage(self.window)  # 进入管理员子菜单操作界面


class NewSell:
    def __init__(self, parent_window):
        self.value1 = None
        self.result = None
        self.row_info = None
        self.row = None
        parent_window.update()
        parent_window.destroy()  # 自动销毁上一个界面
        self.window = Tk()  # 初始框的声明
        self.window.title('管理面板')
        self.window.geometry("800x570+300+30")  # 初始窗口在屏幕中的位置
        self.frame_left_top = tk.Frame(width=800, height=200)  # 指定框架，在窗口上可以显示，这里指定四个框架
        self.frame_right_top = tk.Frame(width=800, height=50)
        self.frame_center = tk.Frame(width=800, height=200)
        self.frame_bottom = tk.Frame(width=800, height=50)
        self.id = []
        self.name = []
        self.model = []
        self.unit = []
        self.num = []
        self.price = []
        self.sum = []
        self.var_id = StringVar()  # 声明学号
        self.var_name = StringVar()  # 声明姓名
        self.var_reminder = StringVar()  # 声明性别
        self.var_model = StringVar()  # 声明年龄
        self.var_unit = StringVar()  # 声明性别
        self.var_cost = StringVar()  # 声明性别
        self.var_price = StringVar()  # 声明年龄
        self.id1 = []
        self.name1 = []
        self.reminder1 = []
        self.model1 = []
        self.unit1 = []
        self.cost1 = []
        self.price1 = []
        self.id11 = 0
        self.var_id1 = StringVar()  # 声明学号
        self.var_name1 = StringVar()  # 声明姓名
        self.var_reminder1 = StringVar()  # 声明性别
        self.var_model1 = StringVar()  # 声明年龄
        self.var_unit1 = StringVar()  # 声明性别
        self.var_cost1 = StringVar()  # 声明性别
        self.var_price1 = StringVar()  # 声明年龄
        self.list()
        self.list1()

        # 定义右上方区域
        self.tree.bind('<Button-1>', self.click)
        # self.right_top_title = Label(self.frame_right_top, text="操作：", font=('微软雅黑', 15))
        self.tree1.bind('<Button-1>', self.click1)  # 左键获取位置(tree.bind可以绑定一系列的事件，可以搜索ttk相关参数查看)
        self.right_top_button4 = ttk.Button(self.frame_right_top, text='售出选中商品',
                                            width=20, command=self.edit_num, padding=5)
        self.right_top_button5 = ttk.Button(self.frame_right_top, text='删除',
                                            width=20, command=self.del_sell, padding=5)
        self.right_top_button6 = ttk.Button(self.frame_right_top, text='完成出货',
                                            width=20, command=self.push_sell, padding=5)

        # 定义下方区域，查询功能块
        self.search = StringVar()
        self.right_bottom_search_entry = Entry(self.frame_bottom, textvariable=self.search, font=('微软雅黑', 15))
        self.right_bottom_button = ttk.Button(self.frame_bottom, text='商品名称查询', width=20, command=self.put_data1,
                                              padding=5)
        self.right_bottom_button.grid(row=0, column=1, padx=20, pady=0)  # 位置设置
        self.right_bottom_search_entry.grid(row=0, column=0, padx=0, pady=0)

        self.right_bottom_button1 = ttk.Button(self.frame_bottom, text='清除查询结果', width=20, command=self.clean_search,
                                               padding=5)
        self.right_bottom_button1.grid(row=0, column=2, padx=20, pady=0)

        # 右上角按钮的位置设置
        self.right_top_button4.grid(row=2, column=0, padx=0, pady=0)
        self.right_top_button5.grid(row=2, column=1, padx=20, pady=0)
        self.right_top_button6.grid(row=2, column=2, padx=20, pady=0)

        # 整体区域定位，利用了Frame和grid进行布局
        self.frame_left_top.grid(row=0, column=0, padx=2, pady=5)
        self.frame_right_top.grid(row=1, column=0, padx=30, pady=30)
        self.frame_center.grid(row=3, column=0, columnspan=2, padx=4, pady=5)
        self.frame_bottom.grid(row=2, column=0, columnspan=2)

        # 设置固定组件，(0)就是将组件进行了固定
        self.frame_left_top.grid_propagate(False)
        self.frame_right_top.grid_propagate(False)
        self.frame_center.grid_propagate(False)
        self.frame_bottom.grid_propagate(False)

        self.frame_left_top.tkraise()  # 开始显示主菜单，tkraise()提高z轴的顺序（不太懂）
        self.frame_right_top.tkraise()  # 开始显示主菜单
        self.frame_center.tkraise()  # 开始显示主菜单
        self.frame_bottom.tkraise()  # 开始显示主菜单

        self.window.protocol("WM_DELETE_WINDOW", self.back)  # 捕捉右上角关闭点击，执行back方法
        self.window.mainloop()  # 进入消息循环

    def push_sell(self):
        td = threading.Thread(target=open_window)
        td.Daemon = True
        td.start()
        path = filedialog.asksaveasfilename(title=u'保存文件', initialfile='导出表单.xlsx',
                                            filetypes=(("Excel files", "*.xlsx"),))
        if not path:
            return
        str1 = path.split(sep=".xlsx", maxsplit=1)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet 1"
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from SELL")
        results = cursor.fetchall()

        self.name = []
        self.model = []
        self.unit = []
        self.num = []
        self.price = []
        self.sum = []
        # 向表格中插入数据
        for row in results:

            self.name.append(row[1])
            self.model.append(row[2])
            self.unit.append(row[3])
            self.num.append(row[4])
            self.price.append(row[5])
            self.sum.append(row[6])
        n = 0
        title = ['名称', '型号', '单位', '数量', '单价', '总价']
        ws.append(title)
        for var in [self.name, self.model, self.unit, self.num, self.price, self.sum]:
            m = var[:]
            n = n + 1
            for i in range(len(m)):
                ws.cell(i + 2, n).value = m[i]
        try:
            wb.save(str1[0] + '.xlsx')
        except PermissionError:
            messagebox.showinfo('警告！', '导出失败，请检查文件是否被占用！')
        else:
            self.push_sale()
            delete_sellout()
            self.put_data()
            messagebox.showinfo('提示！', '导出成功！')

    def push_sale(self):
        i = datetime.datetime.now()
        date1 = "%s/%s/%s %s:%s:%s" % (i.year, i.month, i.day, i.hour, i.minute, i.second)
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        e = conn.cursor().execute("SELECT * FROM SELL")
        result1 = e.fetchall()
        self.value1 = result1[0]
        for i in range(len(result1)):
            self.value1 = result1[i]
            sql1 = "SELECT * FROM COMPANY where ID='%s'" % self.value1[0]
            e = conn.cursor().execute(sql1)
            result2 = e.fetchall()
            value2 = result2[0]
            cost_sum = value2[5] * self.value1[4]
            price = value2[6] * self.value1[4]
            get = price - cost_sum
            sql2 = "INSERT INTO SALE (DATE,ID,NAME,MODEL,UNIT,COST,PRICE,NUM,COST_SUM,SUM,GET) \
                              VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
                   (date1, value2[0], value2[1], value2[3], value2[4], value2[5], value2[6], self.value1[4], cost_sum,
                    price, get)
            try:
                conn.cursor().execute(sql2)  # 执行sql语句
                conn.commit()  # 提交到数据库执行
            except Warning:
                conn.rollback()
                conn.close()
                messagebox.showinfo('警告！', '数据库写入操作失败！')

    def del_sell(self):
        if not self.row_info:
            messagebox.showinfo('警告！', '请点击要删除的项目！')
        else:
            res = messagebox.askyesnocancel('警告！', '是否删除所选数据？')
            if res:
                # 打开数据库连接
                conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
                c = conn.cursor()
                sql = "DELETE from SELL where ID='%s'" % (self.row_info[0])
                try:
                    c.execute(sql)
                    conn.commit()
                except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                    conn.rollback()
                    messagebox.showinfo('警告！', '删除失败，数据库删除操作失败！')
                else:
                    print(self.row_info)
                    self.tree.delete(self.tree.selection()[0])  # 删除所选行
                    sql = "SELECT REMINDER from COMPANY where ID='%s'" % (self.row_info[0])
                    cursor = c.execute(sql)
                    results = cursor.fetchall()
                    m = max(results[0])
                    k = m + int(self.row_info[4])
                    sql = "UPDATE COMPANY set REMINDER = '%s' where ID='%s'" % (k, self.row_info[0])
                    try:
                        c.execute(sql)
                        conn.commit()
                    except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                        messagebox.showinfo('警告！', '删除失败，数据库删除操作失败！')
                    else:
                        conn.close()
                        self.put_data1()
                        messagebox.showinfo('提示！', '删除成功！')
                        self.row_info = []

    def clean_search(self):
        self.search.set('')
        self.put_data1()

    # 将查到的信息放到中间的表格中
    def put_data1(self):
        self.del_button1()  # 先将表格内的内容全部清空
        # 打开数据库连接，准备查找指定的信息
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        d = self.search.get()
        e = conn.cursor().execute("SELECT * FROM COMPANY WHERE name like '%" + d + "%'")
        results = e.fetchall()
        self.id1 = []
        self.name1 = []
        self.reminder1 = []
        self.model1 = []
        self.unit1 = []
        self.cost1 = []
        self.price1 = []
        # 向表格中插入数据
        for row in results:
            self.id1.append(row[0])
            self.name1.append(row[1])
            self.reminder1.append(row[2])
            self.model1.append(row[3])
            self.unit1.append(row[4])
            self.cost1.append(row[5])
            self.price1.append(row[6])

        for i in range(min(len(self.id1), len(self.name1), len(self.reminder1),
                           len(self.model1), len(self.unit1), len(self.cost1), len(self.price1))):  # 写入数据
            self.tree1.insert('', i, values=(self.id1[i], self.name1[i], self.reminder1[i],
                                             self.model1[i], self.unit1[i], self.cost1[i], self.price1[i]))

        for col in self.columns1:  # 绑定函数，使表头可排序
            self.tree1.heading(col, text=col,
                               command=lambda _col=col: self.tree_sort_column(self.tree1, _col, False))

    def put_data(self):
        self.del_button()
        self.id = []
        self.name = []
        self.model = []
        self.unit = []
        self.num = []
        self.price = []
        self.sum = []
        # 打开数据库连接
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from SELL")
        results = cursor.fetchall()
        print(results)
        # 向表格中插入数据
        for row in results:
            self.id.append(row[0])
            self.name.append(row[1])
            self.model.append(row[2])
            self.unit.append(row[3])
            self.num.append(row[4])
            self.price.append(row[5])
            self.sum.append(row[6])

        for i in range(min(len(self.id), len(self.name), len(self.model),
                           len(self.unit), len(self.num), len(self.price), len(self.sum))):  # 写入数据
            self.tree.insert('', i, values=(self.id[i], self.name[i], self.model[i],
                                            self.unit[i], self.num[i], self.price[i], self.sum[i]))

        for col in self.columns:  # 绑定函数，使表头可排序
            self.tree.heading(col, text=col,
                              command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

    # 清空表格中的所有信息
    def del_button(self):
        x = self.tree.get_children()
        for item in x:
            self.tree.delete(item)

    # 在表格上的点击事件，这里是作用就是一点击表格就可以将信息直接写到左上角的框框中
    def click(self, event):
        if not self.tree.identify_row(event.y):
            return
        self.row = self.tree.identify_row(event.y)  # 行
        print(self.row)
        self.row_info = self.tree.item(self.row, "values")
        self.var_id.set(self.row_info[0])
        self.id1 = self.var_id.get()
        self.var_name.set(self.row_info[1])
        self.var_reminder.set(self.row_info[2])
        self.var_model.set(self.row_info[3])
        self.var_unit.set(self.row_info[4])
        self.var_cost.set(self.row_info[5])
        self.var_price.set(self.row_info[6])
        self.right_top_id_entry = Entry(self.frame_left_top, state='disabled', textvariable=self.var_id,
                                        font=('微软雅黑', 15))

    def del_button1(self):
        x = self.tree1.get_children()
        for item in x:
            self.tree1.delete(item)

    # 在表格上的点击事件，这里是作用就是一点击表格就可以将信息直接写到左上角的框框中
    def click1(self, event):
        if not self.tree1.identify_row(event.y):
            return
        self.row1 = self.tree1.identify_row(event.y)  # 行
        print(self.row1)
        self.row_info1 = self.tree1.item(self.row1, "values")
        self.var_id1.set(self.row_info1[0])
        self.id11 = self.var_id1.get()
        self.var_name1.set(self.row_info1[1])
        self.var_reminder1.set(self.row_info1[2])
        self.var_model1.set(self.row_info1[3])
        self.var_unit1.set(self.row_info1[4])
        self.var_cost1.set(self.row_info1[5])
        self.var_price1.set(self.row_info1[6])
        self.right_top_id_entry = Entry(self.frame_left_top, state='disabled', textvariable=self.var_id1,
                                        font=('微软雅黑', 15))

    # 点击中间的表格的表头，可以将那一列进行排序
    def tree_sort_column(self, tv, col, reverse):  # Treeview、列名、排列方式
        m = [(tv.set(k, col), k) for k in tv.get_children('')]
        m.sort(reverse=reverse)  # 排序方式
        for index, (val, k) in enumerate(m):  # 根据排序后索引移动
            tv.move(k, '', index)
        tv.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(tv, _col, not reverse))

    def sellout(self):
        uuu = int(self.row_info1[6])
        price = uuu * self.result
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        sql = "INSERT INTO SELL (ID,NAME,MODEL,UNIT,NUM,PRICE,SUM) \
                                           VALUES ('%s', '%s', '%s', '%s', '%s', '%s', '%s')" % \
              (self.row_info1[0], self.row_info1[1],
               self.row_info1[3], self.row_info1[4], self.result, self.row_info1[6], price)
        try:
            c.execute(sql)  # 执行sql语句
            conn.commit()  # 提交到数据库执行
        except (sqlite3.IntegrityError, sqlite3.ProgrammingError, sqlite3.OperationalError):
            conn.rollback()
            conn.close()
            messagebox.showinfo('警告！', '数据库写入操作失败！')
        self.put_data()
        conn.close()

    def edit_num(self):
        try:
            self.row_info1 = self.tree1.item(self.row1, "values")
        except TclError:
            messagebox.showinfo('警告！', '请选择要修改的项目')
            return FALSE
        td = threading.Thread(target=open_window)
        td.Daemon = True
        td.start()
        self.result = tk.simpledialog.askinteger(title=' ', prompt='售出数量：')
        if not self.result:
            return
        self.id11 = self.var_id1.get()
        d = self.result
        m = self.row_info1[2]
        m1 = int(m)
        if d <= 0:
            messagebox.showinfo('警告！', '请输入大于0的数')
            return FALSE
        k = m1 - d
        if k >= 0:
            # 打开数据库连接
            conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
            c = conn.cursor()
            sql = "UPDATE COMPANY set REMINDER = '%s' where ID='%s'" % \
                  (k, self.id11)
            try:
                c.execute(sql)
                conn.commit()
            except (sqlite3.ProgrammingError, sqlite3.OperationalError):
                conn.rollback()
                conn.close()
                messagebox.showinfo('警告！', '更新失败，数据库写入操作失败！')
            else:
                conn.close()
                self.tree1.item(self.tree1.selection()[0], values=(
                    self.var_id1.get(), self.var_name1.get(), k,
                    self.var_model1.get(), self.var_unit1.get(), self.var_cost1.get(), self.var_price1.get()))
                k1 = str(k)
                self.var_reminder1.set(k1)
                self.sellout()
                messagebox.showinfo('提示！', '更新成功！')
        else:
            messagebox.showinfo('警告！', '更新失败，余货量不能为负数！')

    def list(self):
        # 定义下方中心列表区域
        self.columns = ("商品id", "名称", "型号", "单位", "数量", "单价", "总价")
        self.tree = ttk.Treeview(self.frame_left_top, show="headings", height=8, columns=self.columns)
        # 添加竖直滚动条
        self.vbar = ttk.Scrollbar(self.frame_left_top, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree.configure(yscrollcommand=self.vbar.set)

        # 定义id1为修改id时的暂存变量，这个是为了更新信息而设计的
        self.id1 = 0
        # 表格的标题
        self.tree.column("商品id", width=100, anchor='center')
        self.tree.column("名称", width=100, anchor='center')
        self.tree.column("型号", width=150, anchor='center')
        self.tree.column("单位", width=100, anchor='center')
        self.tree.column("数量", width=100, anchor='center')
        self.tree.column("单价", width=100, anchor='center')
        self.tree.column("总价", width=100, anchor='center')

        self.var_id = StringVar()  # 声明学号
        self.var_name = StringVar()  # 声明姓名
        self.var_reminder = StringVar()  # 声明性别
        self.var_model = StringVar()  # 声明年龄
        self.var_unit = StringVar()  # 声明性别
        self.var_cost = StringVar()  # 声明性别
        self.var_price = StringVar()  # 声明年龄
        # grid方法将tree和vbar进行布局
        self.tree.grid(row=0, column=0, sticky=NSEW)
        self.vbar.grid(row=0, column=1, sticky=NS)

        # 定义几个数组，为中间的那个大表格做一些准备

        self.id = []
        self.name = []
        self.model = []
        self.unit = []
        self.num = []
        self.price = []
        self.sum = []
        # 打开数据库连接
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from SELL")
        results = cursor.fetchall()
        print(results)
        # 向表格中插入数据
        for row in results:
            self.id.append(row[0])
            self.name.append(row[1])
            self.model.append(row[2])
            self.unit.append(row[3])
            self.num.append(row[4])
            self.price.append(row[5])
            self.sum.append(row[6])

        for i in range(min(len(self.id), len(self.name), len(self.model),
                           len(self.unit), len(self.num), len(self.price), len(self.sum))):  # 写入数据
            self.tree.insert('', i, values=(self.id[i], self.name[i], self.model[i],
                                            self.unit[i], self.num[i], self.price[i], self.sum[i]))

        # 绑定函数，使表头可排序
        for col in self.columns:
            self.tree.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(self.tree, _col, False))

    def list1(self):
        # 定义下方中心列表区域

        self.columns1 = ("商品id", "名称", "余货量", "型号", "单位", "成本价", "单价")
        self.tree1 = ttk.Treeview(self.frame_center, show="headings", height=8, columns=self.columns1)
        # 添加竖直滚动条
        self.vbar1 = ttk.Scrollbar(self.frame_center, orient=VERTICAL, command=self.tree.yview)
        # 定义树形结构与滚动条
        self.tree1.configure(yscrollcommand=self.vbar1.set)

        # 定义id1为修改id时的暂存变量，这个是为了更新信息而设计的
        self.id1 = 0
        # 表格的标题
        self.tree1.column("商品id", width=100, anchor='center')
        self.tree1.column("名称", width=150, anchor='center')
        self.tree1.column("余货量", width=100, anchor='center')
        self.tree1.column("型号", width=100, anchor='center')
        self.tree1.column("单位", width=100, anchor='center')
        self.tree1.column("成本价", width=100, anchor='center')
        self.tree1.column("单价", width=100, anchor='center')

        # grid方法将tree和vbar进行布局
        self.tree1.grid(row=0, column=0, sticky=NSEW)
        self.vbar1.grid(row=0, column=1, sticky=NS)

        # 定义几个数组，为中间的那个大表格做一些准备

        self.id1 = []
        self.name1 = []
        self.reminder1 = []
        self.model1 = []
        self.unit1 = []
        self.cost1 = []
        self.price1 = []
        # 打开数据库连接
        conn = sqlite3.connect(os.path.expanduser("~/sql/list.db"))
        c = conn.cursor()
        cursor = c.execute("SELECT *  from COMPANY")
        results = cursor.fetchall()
        print(results)
        # 向表格中插入数据
        for row in results:
            self.id1.append(row[0])
            self.name1.append(row[1])
            self.reminder1.append(row[2])
            self.model1.append(row[3])
            self.unit1.append(row[4])
            self.cost1.append(row[5])
            self.price1.append(row[6])

        for i in range(min(len(self.id1), len(self.name1), len(self.reminder1),
                           len(self.model1), len(self.unit1), len(self.cost1), len(self.price1))):  # 写入数据
            self.tree1.insert('', i, values=(self.id1[i], self.name1[i], self.reminder1[i],
                                             self.model1[i], self.unit1[i], self.cost1[i], self.price1[i]))

        # 绑定函数，使表头可排序
        for col in self.columns1:
            self.tree1.heading(col, text=col, command=lambda _col=col: self.tree_sort_column(self.tree1, _col, False))

    def back(self):
        AdminPage(self.window)  # 进入管理员子菜单操作界面


if __name__ == '__main__':
    # 实例化Application
    window = tk.Tk()
    AdminPage(window)
    # SellOut(window)
